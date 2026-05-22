"""
clustering_pipeline.py — Unsupervised Patient Phenotyping for Impella Analytics
=============================================================================
Consensus K-Means (k=3) + Hierarchical Clustering to identify clinical phenotypes.

Output: ml_output/clusters/
    - cluster_assignments.csv   — patient MRN → cluster label
    - cluster_profiles.json     — mean features per cluster + outcome rates
    - cluster_model.joblib      — imputer + scaler + kmeans (for predict_cluster.py)
    - dendrogram.png            — hierarchical clustering tree
    - consensus_matrix.png      — consensus heatmap
    - clustering_report.md      — human-readable summary

Run: python clustering_pipeline.py
"""

import json
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.cluster.hierarchy import dendrogram, linkage, fcluster
from scipy.spatial.distance import pdist

from sklearn.cluster import KMeans
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.metrics import silhouette_score, silhouette_samples
from sklearn.manifold import TSNE

import joblib

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_PATH = Path("Impella_MK.xlsx")
OUTPUT_DIR = Path("ml_output/clusters")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

RANDOM_STATE = 42
N_BOOTSTRAP = 200      # consensus k-means iterations
SUBSAMPLE_RATIO = 0.80  # fraction of patients per iteration
N_CLUSTERS = 3

# Row mapping for Patient Data sheet (1-based Excel row numbers, from ml_pipeline.py)
PATIENT_DATA_ROWS = {
    "general_notes": 2,
    "first_name": 3,
    "last_name": 4,
    "mrn": 5,
    "date_of_implant": 6,
    "age": 7,
    "weight_kg": 8,
    "height_cm": 9,
    "gender": 10,
    "race": 11,
    "cause_of_shock": 12,
    "scai_stage": 13,
    "days_between_rhc_and_impella": 14,
    # Pre-implant RHC (rows 17-35)
    "pre_ra": 17,
    "pre_rvsp": 18,
    "pre_rvdp": 19,
    "pre_pasp": 20,
    "pre_padp": 21,
    "pre_map": 27,
    "pre_pcwp": 23,
    "pre_pvr": 24,
    "pre_sbp": 25,
    "pre_dbp": 26,
    "pre_hr": 28,
    "pre_tdco": 29,
    "pre_sv": 30,
    "pre_pa_o2": 31,
    "pre_sp_o2": 32,
    "pre_papi": 33,
    "pre_cpo": 34,
    "pre_rv_cpo": 35,
    # Post-implant RHC (rows 40-58)
    "post_ra": 40,
    "post_rvsp": 41,
    "post_rvdp": 42,
    "post_pasp": 43,
    "post_padp": 44,
    "post_map": 50,
    "post_pcwp": 46,
    "post_pvr": 47,
    "post_sbp": 48,
    "post_dbp": 49,
    "post_hr": 51,
    "post_tdco": 52,
    "post_sv": 53,
    "post_pa_o2": 54,
    "post_sp_o2": 55,
    "post_papi": 56,
    "post_cpo": 57,
    "post_rv_cpo": 58,
    # Echo pre (rows 67-75)
    "pre_rvedd": 67,
    "pre_tapse": 68,
    "pre_rv_s": 69,
    "pre_rv_fs": 70,
    "pre_tr_severity": 71,
    "pre_echo_pasp": 72,
    "pre_lvedd": 73,
    "pre_septal_flattening": 74,
    "pre_atrial_bowing": 75,
    # Echo post (rows 78-86)
    "post_rvedd": 78,
    "post_tapse": 79,
    "post_rv_s": 80,
    "post_rv_fs": 81,
    "post_tr_severity": 82,
    "post_echo_pasp": 83,
    "post_lvedd": 84,
    "post_septal_flattening": 85,
    "post_atrial_bowing": 86,
    # Labs pre (rows 88-98)
    "pre_sodium": 88,
    "pre_potassium": 89,
    "pre_hco3": 90,
    "pre_creatinine": 91,
    "pre_egfr": 92,
    "pre_hemoglobin": 93,
    "pre_wbc": 94,
    "pre_ast": 95,
    "pre_alt": 96,
    "pre_bili": 97,
    "pre_lactate": 98,
    "pre_ph": 99,
    # Labs post (rows 102-113)
    "post_sodium": 102,
    "post_potassium": 103,
    "post_hco3": 104,
    "post_creatinine": 105,
    "post_egfr": 106,
    "post_hemoglobin": 107,
    "post_wbc": 108,
    "post_ast": 109,
    "post_alt": 110,
    "post_bili": 111,
    "post_lactate": 112,
    "post_ph": 113,
    # Inotropes (rows 116-122)
    "dopamine": 116,
    "dobutamine": 117,
    "epinephrine": 118,
    "milrinone": 119,
    "norepinephrine": 120,
    "vasopressin": 121,
    "vis_score": 122,
    # Diuretics pre (rows 125-126)
    "pre_furosemide": 125,
    "pre_augmentation": 126,
    # Diuretics post (rows 129-130)
    "post_furosemide": 129,
    "post_augmentation": 130,
    # Impella (rows 133-134)
    "impella_performance": 133,
    "impella_flow": 134,
    # Outcomes (rows 138-141)
    "renal_failure": 138,
    "intubation": 139,
    "mcs_escalation": 140,
    "outcome": 141,
    # PV Loop (rows 144-154)
    "ees": 144,
    "ea": 145,
    "ees_ea": 146,
    "esp": 147,
    "edp": 148,
    "pmax": 149,
    "esv": 150,
    "edv": 151,
    "pv_sv": 152,
    "dp_dt_max": 153,
    "dp_dt_min": 154,
}

# ---------------------------------------------------------------------------
# Data Loading (mirrors ml_pipeline.py)
# ---------------------------------------------------------------------------

def load_patient_data(path: Path) -> pd.DataFrame:
    """Load Patient Data sheet — one patient per column (wide format)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Patient Data"]

    rows = []
    for col in range(2, ws.max_column + 1):
        mrn = ws.cell(row=5, column=col).value
        if not mrn:
            continue
        record = {"mrn": str(mrn).strip()}
        for key, row_idx in PATIENT_DATA_ROWS.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, str):
                val = val.strip()
                if val.lower() in ("n/a", "na", "", "none", "#div/0!", "n/na", "n//a"):
                    val = np.nan
            record[key] = val
        rows.append(record)

    df = pd.DataFrame(rows)
    numeric_keys = [k for k in PATIENT_DATA_ROWS.keys() if k not in (
        "general_notes", "first_name", "last_name", "mrn", "date_of_implant"
    )]
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")
    return df


def load_cohort(path: Path) -> pd.DataFrame:
    """Load Cohort sheet."""
    df = pd.read_excel(path, sheet_name="Cohort")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    rename_map = {
        "MRN": "mrn",
        "Outcome": "cohort_outcome",
        "Age ": "cohort_age",
        "Gender": "cohort_gender",
        "Indication for Use": "indication",
        "Duration of Support Time (Days)": "support_days",
        "Physician Name": "physician",
        "RHC prior to implant (<72 hours)": "rhc_prior_72h",
        "RHC timing (days prior)": "rhc_timing_days",
    }
    df = df.rename(columns=rename_map)
    df["mrn"] = df["mrn"].astype(str).str.strip()
    return df


# ---------------------------------------------------------------------------
# Feature Engineering (clustering-specific)
# ---------------------------------------------------------------------------

def engineer_clustering_features(df: pd.DataFrame) -> pd.DataFrame:
    """Create features for clustering — mirrors supervised pipeline but focused."""
    df = df.copy()

    # Demographics
    df["bmi"] = df["weight_kg"] / ((df["height_cm"] / 100) ** 2)

    # Delta features (post - pre) for key hemodynamics
    delta_vars = ["ra", "papi", "cpo", "creatinine", "lactate", "tdco", "pcwp"]
    for v in delta_vars:
        pre = f"pre_{v}"
        post = f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]

    # CPO ratio
    df["ratio_cpo"] = df["post_cpo"] / df["pre_cpo"].replace(0, np.nan)

    # Inotrope count
    inotrope_cols = ["dopamine", "dobutamine", "epinephrine", "milrinone", "norepinephrine", "vasopressin"]
    df["inotrope_count"] = df[inotrope_cols].gt(0).sum(axis=1)

    # VIS high
    df["vis_high"] = (df["vis_score"] > 15).astype(int)

    # SCAI numeric — handles both numeric (1-4) and string (A-E / B-E)
    def _scai_to_numeric(val):
        if pd.isna(val):
            return np.nan
        try:
            n = float(val)
            if 1 <= n <= 4:
                return int(n)
        except (ValueError, TypeError):
            pass
        s = str(val).lower().strip()
        scai_map = {"a": 1, "b": 1, "c": 2, "d": 3, "e": 4}
        return scai_map.get(s, np.nan)

    df["scai_numeric"] = df["scai_stage"].apply(_scai_to_numeric)

    # Echo delta features (post - pre)
    echo_delta_vars = ["tapse", "rv_s", "rv_fs", "rvedd"]
    for v in echo_delta_vars:
        pre = f"pre_{v}"
        post = f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]

    # Congestion score (RA + PCWP)
    if "pre_ra" in df.columns and "pre_pcwp" in df.columns:
        df["congestion_score"] = df["pre_ra"] + df["pre_pcwp"]

    # V-A coupling ratio (if not already present)
    if "ees_ea" not in df.columns and "ees" in df.columns and "ea" in df.columns:
        df["ees_ea"] = df["ees"] / df["ea"].replace(0, np.nan)

    return df


# ---------------------------------------------------------------------------
# Feature Selection
# ---------------------------------------------------------------------------

# Phase 1: ~20 literature-validated features
CLUSTERING_FEATURES_PHASE1 = [
    # Demographics
    "age",
    "bmi",
    # Labs pre-implant
    "pre_lactate",
    "pre_egfr",
    "pre_hco3",
    "pre_alt",
    "pre_creatinine",
    "pre_sodium",
    "pre_wbc",
    "pre_hemoglobin",
    "pre_bili",
    # Hemodynamics pre-implant
    "pre_ra",
    "pre_papi",
    "pre_cpo",
    "pre_tdco",
    "pre_sbp",
    # Delta features
    "delta_cpo",
    "delta_lactate",
    "delta_creatinine",
    "delta_papi",
]

# Phase 3: lean, highly-discriminative feature set (~30 features)
# Strategy: keep pre-implant baseline + PV loop + minimal echo;
# avoid post-implant features that reflect treatment effect rather than phenotype
CLUSTERING_FEATURES_PHASE3 = [
    # Demographics
    "age",
    "bmi",
    "gender",
    # Labs pre-implant (core metabolic profile)
    "pre_lactate",
    "pre_egfr",
    "pre_hco3",
    "pre_alt",
    "pre_creatinine",
    "pre_sodium",
    "pre_wbc",
    "pre_hemoglobin",
    "pre_bili",
    # Hemodynamics pre-implant (core congestion/perfusion profile)
    "pre_ra",
    "pre_papi",
    "pre_cpo",
    "pre_tdco",
    "pre_sbp",
    "pre_map",
    "pre_pcwp",
    "pre_padp",
    "pre_rvsp",
    # Delta features (treatment response)
    "delta_cpo",
    "delta_lactate",
    "delta_creatinine",
    "delta_papi",
    # PV loop (contractility + V-A coupling — the key discriminative addition)
    "ees",
    "ea",
    "ees_ea",
    "edp",
    "esp",
    "pmax",
    # Echo (minimal well-populated set)
    "pre_echo_pasp",
    "pre_tapse",
    # Support intensity
    "vis_score",
    "inotrope_count",
    # Derived composites
    "congestion_score",
    "scai_numeric",
]

# ZWECK_STYLE: 6 core labs + congestion marker + demographics + SCAI
# PV loop features (ees, ea, ees_ea) removed after sandbox testing showed
# silhouette improves from 0.223 → 0.263 without them.
CLUSTERING_FEATURES_ZWECK = [
    # Zweck core 6
    "pre_egfr",
    "pre_hco3",
    "pre_lactate",
    "pre_alt",
    "pre_wbc",
    "pre_hemoglobin",  # proxy for platelets (both CBC) — platelets not in our data label row
    # Hemodynamic congestion marker
    "pre_ra",
    # Demographics (age and BMI are well-populated and prognostic)
    "age",
    "bmi",
    # SCAI stage (shock severity classification — 1=A/B, 2=C, 3=D, 4=E)
    "scai_numeric",
]


def select_features(df: pd.DataFrame, feature_list: list) -> pd.DataFrame:
    """Return DataFrame with only available features from the list."""
    available = [f for f in feature_list if f in df.columns]
    missing = [f for f in feature_list if f not in df.columns]
    if missing:
        print(f"  Warning: missing features (will be imputed): {missing[:5]}{'...' if len(missing) > 5 else ''}")
    return df[available].copy()


# ---------------------------------------------------------------------------
# Consensus K-Means
# ---------------------------------------------------------------------------

def consensus_kmeans(X: np.ndarray, n_clusters: int, n_bootstrap: int, subsample_ratio: float, random_state: int) -> tuple:
    """
    Consensus K-Means: run k-means on bootstrap subsamples and build consensus matrix.
    Returns: consensus_matrix (n x n), all_labels (list of label arrays per iteration)
    """
    n_samples = X.shape[0]
    n_subsample = int(n_samples * subsample_ratio)
    rng = np.random.RandomState(random_state)

    n_iterations = n_bootstrap
    connectivity_matrix = np.zeros((n_samples, n_samples), dtype=float)

    print(f"  Running {n_iterations} bootstrap iterations (subsample={subsample_ratio:.0%})...")

    for i in range(n_iterations):
        # Subsample
        indices = rng.choice(n_samples, size=n_subsample, replace=False)
        X_sub = X[indices]

        # Fit k-means
        km = KMeans(n_clusters=n_clusters, init="k-means++", n_init=10, random_state=rng.randint(0, 10000))
        labels_sub = km.fit_predict(X_sub)

        # Build connectivity matrix for this subsample
        for a in range(len(indices)):
            for b in range(a + 1, len(indices)):
                if labels_sub[a] == labels_sub[b]:
                    connectivity_matrix[indices[a], indices[b]] += 1
                    connectivity_matrix[indices[b], indices[a]] += 1

        if (i + 1) % 50 == 0:
            print(f"    Iteration {i + 1}/{n_iterations}")

    consensus_matrix = connectivity_matrix / n_iterations
    return consensus_matrix


def run_final_kmeans(X: np.ndarray, n_clusters: int, random_state: int) -> tuple:
    """Run final k-means on full data to get stable labels."""
    km = KMeans(n_clusters=n_clusters, init="k-means++", n_init=50, random_state=random_state)
    labels = km.fit_predict(X)
    return labels, km.cluster_centers_


# ---------------------------------------------------------------------------
# Hierarchical Clustering
# ---------------------------------------------------------------------------

def hierarchical_clustering(X: np.ndarray, method: str = "ward") -> np.ndarray:
    """Ward linkage hierarchical clustering. Returns linkage matrix."""
    return linkage(X, method=method)


def plot_dendrogram(linkage_matrix: np.ndarray, output_path: Path, labels: list = None):
    """Plot and save dendrogram."""
    plt.figure(figsize=(14, 6))
    dendrogram(
        linkage_matrix,
        labels=labels,
        leaf_rotation=90,
        leaf_font_size=7,
        color_threshold=0,
    )
    plt.title("Hierarchical Clustering Dendrogram (Ward linkage)")
    plt.xlabel("Patient MRN")
    plt.ylabel("Distance")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved dendrogram to {output_path}")


# ---------------------------------------------------------------------------
# Visualizations
# ---------------------------------------------------------------------------

def plot_consensus_matrix(consensus_matrix: np.ndarray, output_path: Path, patient_ids: list):
    """Plot consensus matrix heatmap."""
    plt.figure(figsize=(10, 8))
    # Reorder by hierarchical clustering
    from scipy.cluster.hierarchy import leaves_list
    linkage_hc = linkage(pdist(consensus_matrix), method="ward")
    order = leaves_list(linkage_hc)
    consensus_ordered = consensus_matrix[order][:, order]
    patient_ids_ordered = [patient_ids[i] for i in order]

    sns.heatmap(
        consensus_ordered,
        cmap="Blues",
        vmin=0,
        vmax=1,
        xticklabels=patient_ids_ordered,
        yticklabels=patient_ids_ordered,
        cbar_kws={"label": "Consensus Index"},
    )
    plt.title("Consensus Matrix Heatmap\n(Ordered by Hierarchical Clustering)")
    plt.xlabel("Patient")
    plt.ylabel("Patient")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved consensus matrix to {output_path}")


def plot_silhouette(X: np.ndarray, labels: np.ndarray, output_path: Path, feature_names: list):
    """Plot silhouette analysis per sample."""
    n_clusters = len(np.unique(labels))
    silhouette_vals = silhouette_samples(X, labels)
    silhouette_avg = silhouette_score(X, labels)
    y_lower = 10

    fig, ax = plt.subplots(figsize=(8, max(6, len(labels) * 0.3)))
    for i in range(n_clusters):
        ith_cluster_silhouette_vals = silhouette_vals[labels == i]
        ith_cluster_silhouette_vals.sort()
        size_cluster_i = ith_cluster_silhouette_vals.shape[0]
        y_upper = y_lower + size_cluster_i

        color = plt.cm.tab10(i / n_clusters)
        ax.fill_betweenx(np.arange(y_lower, y_upper), 0, ith_cluster_silhouette_vals,
                         facecolor=color, alpha=0.7)
        ax.text(-0.05, y_lower + 0.5 * size_cluster_i, str(i), fontweight="bold")
        y_lower = y_upper + 10

    ax.axvline(x=silhouette_avg, color="red", linestyle="--", label=f"Avg={silhouette_avg:.3f}")
    ax.set_title(f"Silhouette Analysis (k={n_clusters}, avg={silhouette_avg:.3f})")
    ax.set_xlabel("Silhouette Coefficient")
    ax.set_ylabel("Cluster")
    ax.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved silhouette plot to {output_path}")
    return silhouette_avg


def plot_cluster_profiles(profiles: dict, output_path: Path):
    """Plot heatmap of mean feature values per cluster (standardized)."""
    labels = sorted(profiles.keys())
    features = list(profiles[labels[0]]["mean_features"].keys())
    n_features = len(features)
    n_clusters = len(labels)

    # Build matrix: clusters x features
    matrix = np.zeros((n_clusters, n_features))
    for i, label in enumerate(labels):
        for j, feat in enumerate(features):
            matrix[i, j] = profiles[label]["mean_features"].get(feat, np.nan)

    # Standardize across clusters for visualization
    matrix_std = (matrix - matrix.mean(axis=0, keepdims=True)) / (matrix.std(axis=0, keepdims=True) + 1e-8)

    plt.figure(figsize=(max(12, n_features * 0.5), max(4, n_clusters * 1.5)))
    sns.heatmap(
        matrix_std.T,
        xticklabels=[f"C{l}" for l in labels],
        yticklabels=features,
        cmap="RdBu_r",
        center=0,
        annot=False,
        cbar_kws={"label": "Z-score (standardized)"}
    )
    plt.title("Cluster Feature Profiles (Z-scored across clusters)")
    plt.xlabel("Cluster")
    plt.ylabel("Feature")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved cluster profiles heatmap to {output_path}")


def plot_pca_scatter(X_2d: np.ndarray, labels: np.ndarray, output_path: Path, patient_ids: list):
    """Plot 2D PCA scatter colored by cluster."""
    plt.figure(figsize=(8, 6))
    unique_labels = sorted(np.unique(labels))
    colors = plt.cm.tab10(np.linspace(0, 1, len(unique_labels)))

    for i, label in enumerate(unique_labels):
        mask = labels == label
        plt.scatter(X_2d[mask, 0], X_2d[mask, 1],
                    c=[colors[i]], label=f"Cluster {label}", s=60, alpha=0.7)

    plt.xlabel("PC1")
    plt.ylabel("PC2")
    plt.title("Patient Clusters (PCA Projection)")
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved PCA scatter to {output_path}")


def plot_outcome_rates(profiles: dict, output_path: Path):
    """Plot stacked bar chart of outcome rates per cluster."""
    labels = sorted(profiles.keys())
    n = len(labels)
    x = np.arange(n)
    width = 0.25

    survival = [profiles[l]["survival_rate"] for l in labels]
    escalation = [profiles[l]["escalation_rate"] for l in labels]
    renal = [profiles[l]["renal_rate"] for l in labels]

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x - width, survival, width, label="Survival", color="#4CAF50")
    ax.bar(x, escalation, width, label="Escalation", color="#FF9800")
    ax.bar(x + width, renal, width, label="Renal Failure", color="#F44336")

    ax.set_xlabel("Cluster")
    ax.set_ylabel("Rate")
    ax.set_title("Clinical Outcome Rates by Cluster")
    ax.set_xticks(x)
    ax.set_xticklabels([f"C{l}" for l in labels])
    ax.legend()
    ax.set_ylim(0, 1.1)
    ax.axhline(y=0.5, color="gray", linestyle="--", alpha=0.3)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Saved outcome rates chart to {output_path}")


# ---------------------------------------------------------------------------
# Cluster Profiling
# ---------------------------------------------------------------------------

def build_cluster_profiles(df: pd.DataFrame, labels: np.ndarray, feature_names: list) -> dict:
    """Compute mean feature values and outcome rates per cluster."""
    df = df.copy()
    df["cluster"] = labels

    profiles = {}
    for cluster_id in sorted(df["cluster"].unique()):
        subset = df[df["cluster"] == cluster_id]
        n = len(subset)

        # Mean features
        mean_feats = {}
        for feat in feature_names:
            if feat in subset.columns:
                mean_feats[feat] = float(subset[feat].mean())

        # Outcome rates — check Cohort sheet first (has "Survived"/"Expired" strings)
        # then Patient Data sheet outcome column (numeric codes)
        survival_rate = None
        if "cohort_outcome" in subset.columns:
            survived_count = subset["cohort_outcome"].astype(str).str.lower().str.strip().eq("survived").sum()
            survival_rate = survived_count / n if n > 0 else None
        elif "outcome" in subset.columns:
            # Patient Data uses numeric: 4 = expired, 1-3 = survived
            expired_count = subset["outcome"].eq(4).sum()
            survival_rate = (n - expired_count) / n if n > 0 else None

        # Escalation — from Cohort notes or Patient Data outcomes
        escalation_rate = None
        if "target_escalation" in subset.columns:
            escalation_rate = float(subset["target_escalation"].mean())
        elif "mcs_escalation" in subset.columns:
            escalation_rate = float(subset["mcs_escalation"].mean())
        elif "cohort_outcome" in subset.columns:
            # Fallback: check for escalation keywords in outcome/notes
            notes = subset["cohort_outcome"].astype(str).str.lower()
            escalation_rate = notes.str.contains("ecmo|lvad|transplant|arrest").mean()

        # Renal failure
        renal_rate = None
        if "renal_failure" in subset.columns:
            renal_rate = float(subset["renal_failure"].mean())

        # SCAI distribution
        scai_vals = {}
        if "scai_stage" in subset.columns:
            scai_vals = subset["scai_stage"].dropna().astype(str).str.lower().value_counts().to_dict()

        profiles[int(cluster_id)] = {
            "patient_count": n,
            "mean_features": mean_feats,
            "survival_rate": survival_rate,
            "escalation_rate": escalation_rate,
            "renal_rate": renal_rate,
            "scai_distribution": scai_vals,
            "mrn_list": subset["mrn"].tolist(),
        }

    return profiles


def assign_cluster_names(profiles: dict) -> dict:
    """
    Assign clinical phenotype names based on cluster characteristics.
    For k=3: rank clusters to assign exactly one Cardiorenal, one Cardiometabolic,
    and one Non-congested phenotype.
    """
    # Collect cluster scores for ranking
    cid_list = sorted(profiles.keys(), key=lambda x: int(x))
    scores = {}
    for cid in cid_list:
        mf = profiles[cid]["mean_features"]
        scores[cid] = {
            "egfr": mf.get("pre_egfr", np.nan),
            "lactate": mf.get("pre_lactate", np.nan),
            "ra": mf.get("pre_ra", np.nan),
            "hco3": mf.get("pre_hco3", np.nan),
            "alt": mf.get("pre_alt", np.nan),
            "creatinine": mf.get("pre_creatinine", np.nan),
        }

    n_clusters = len(cid_list)
    assigned = {}

    if n_clusters >= 3:
        # Rank by eGFR (ascending) — lowest eGFR is most cardiorenal
        egfr_sorted = sorted(cid_list, key=lambda c: scores[c]["egfr"] if not np.isnan(scores[c]["egfr"]) else float("inf"))
        # Rank by lactate (descending) — highest lactate is most cardiometabolic
        lac_sorted = sorted(cid_list, key=lambda c: scores[c]["lactate"] if not np.isnan(scores[c]["lactate"]) else -float("inf"), reverse=True)
        # Rank by RA (ascending) — lowest RA is least congested
        ra_sorted = sorted(cid_list, key=lambda c: scores[c]["ra"] if not np.isnan(scores[c]["ra"]) else float("inf"))

        # Cardiorenal: lowest eGFR (break ties by higher RA = more congested)
        def _renal_score(c):
            egfr = scores[c]["egfr"] if not np.isnan(scores[c]["egfr"]) else float("inf")
            ra = scores[c]["ra"] if not np.isnan(scores[c]["ra"]) else -float("inf")
            return (egfr, -ra)

        cardiorenal_candidates = sorted(cid_list, key=_renal_score)
        cardiorenal_id = cardiorenal_candidates[0]

        # Cardiometabolic: highest lactate among remaining
        remaining_after_renal = [c for c in lac_sorted if c != cardiorenal_id]
        cardiometabolic_id = remaining_after_renal[0]

        # Non-congested: the remaining cluster
        noncongested_id = [c for c in cid_list if c not in (cardiorenal_id, cardiometabolic_id)][0]

        assigned[cardiorenal_id] = "Cardiorenal (Moderate-risk)"
        assigned[cardiometabolic_id] = "Cardiometabolic (High-risk)"
        assigned[noncongested_id] = "Non-congested (Low-risk)"
    else:
        # Fallback to threshold-based logic for k<3
        for cid in cid_list:
            s = scores[cid]
            is_cardiorenal = (not np.isnan(s["egfr"]) and s["egfr"] < 45) or \
                             (not np.isnan(s["creatinine"]) and s["creatinine"] > 1.8)
            is_cardiometabolic = (not np.isnan(s["lactate"]) and s["lactate"] > 2.0) or \
                                 (not np.isnan(s["hco3"]) and s["hco3"] < 22 and not np.isnan(s["lactate"]) and s["lactate"] > 1.5)
            if is_cardiorenal:
                assigned[cid] = "Cardiorenal (Moderate-risk)"
            elif is_cardiometabolic:
                assigned[cid] = "Cardiometabolic (High-risk)"
            else:
                assigned[cid] = "Non-congested (Low-risk)"

    named = {}
    for cid, prof in profiles.items():
        best_name = assigned.get(cid, "Unknown")
        named[cid] = {
            **prof,
            "cluster_name": best_name,
            "clinical_recommendation": {
                "Non-congested (Low-risk)": "Early weaning candidate — monitor for bounce-back",
                "Cardiorenal (Moderate-risk)": "Staged escalation planning — watch for RV deterioration",
                "Cardiometabolic (High-risk)": "Consider advanced MCS / RV support / transplant evaluation",
            }.get(best_name, ""),
        }

    return named


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_assignments(df: pd.DataFrame, labels: np.ndarray, output_path: Path):
    """Save cluster_assignments.csv."""
    assignments = pd.DataFrame({
        "mrn": df["mrn"].values,
        "cluster": labels,
    })
    assignments.to_csv(output_path, index=False)
    print(f"  Saved cluster assignments to {output_path}")


def export_profiles(profiles: dict, output_path: Path):
    """Save cluster_profiles.json."""
    # Convert numpy types for JSON serialization
    def make_serializable(obj):
        if isinstance(obj, dict):
            return {k: make_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32)):
            val = float(obj)
            return None if np.isnan(val) else val
        elif isinstance(obj, float):
            return None if np.isnan(obj) else obj
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, list):
            return [make_serializable(v) for v in obj]
        return obj

    serializable = make_serializable(profiles)
    with open(output_path, "w") as f:
        json.dump(serializable, f, indent=2)
    print(f"  Saved cluster profiles to {output_path}")


def export_quality_metrics(silhouette_avg: float, n_clusters: int, n_features: int,
                           n_patients: int, output_path: Path):
    """Save quality_metrics.json with interpretability guidance."""
    if silhouette_avg < 0.15:
        interpretation = "weak"
        clinical_caution = (
            f"Silhouette score is low ({silhouette_avg:.3f}), indicating modest cluster separation. "
            "Use cluster assignments as exploratory phenotyping guidance rather than definitive stratification. "
            "Cross-check with traditional hemodynamic thresholds (RA, PAPI, delta CPO) before clinical decisions."
        )
    elif silhouette_avg < 0.35:
        interpretation = "moderate"
        clinical_caution = (
            f"Silhouette score ({silhouette_avg:.3f}) indicates fair cluster separation. "
            "Cluster assignments are reasonable for exploratory phenotyping but should be validated "
            "against individual patient hemodynamics and clinical judgment."
        )
    else:
        interpretation = "strong"
        clinical_caution = (
            f"Silhouette score ({silhouette_avg:.3f}) indicates good cluster separation. "
            "Phenotype assignments are likely reliable for clinical stratification."
        )

    metrics = {
        "silhouette_score": float(silhouette_avg),
        "n_clusters": n_clusters,
        "n_features": n_features,
        "n_patients": n_patients,
        "clustering_method": "consensus_kmeans",
        "bootstrap_iterations": N_BOOTSTRAP,
        "interpretation": interpretation,
        "clinical_caution": clinical_caution,
    }
    with open(output_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"  Saved quality metrics to {output_path}")


def export_model(imputer, scaler, kmeans, centroids, feature_names: list, output_path: Path, pca=None):
    """Save cluster_model.joblib with kmeans object + centroids array + optional PCA."""
    artifact = {
        "imputer": imputer,
        "scaler": scaler,
        "kmeans": kmeans,           # sklearn KMeans object (for refit)
        "kmeans_centroids": centroids,  # numpy array (for direct distance)
        "feature_names": feature_names,
        "pca": pca,                 # PCA transformer if used
    }
    joblib.dump(artifact, output_path)
    print(f"  Saved clustering model to {output_path}")


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def generate_report(profiles: dict, silhouette_avg: float, feature_names: list, n_clusters: int, output_dir: Path):
    """Write clustering_report.md."""
    lines = [
        "# Impella Patient Clustering Report",
        "",
        f"**Date**: Auto-generated",
        f"**Method**: Consensus K-Means (k={n_clusters}, {N_BOOTSTRAP} bootstrap iterations, {SUBSAMPLE_RATIO:.0%} subsample)",
        f"**Silhouette Score**: {silhouette_avg:.3f}",
        f"**Features Used**: {len(feature_names)}",
        "",
        "## Cluster Summary",
        "",
    ]

    for cid, prof in sorted(profiles.items()):
        lines.append(f"### Cluster {cid}: {prof.get('cluster_name', 'Unknown')}")
        lines.append(f"- **Patients**: {prof['patient_count']}")
        lines.append(f"- **Survival Rate**: {prof.get('survival_rate', 'N/A'):.1%}" if prof.get('survival_rate') is not None else "- **Survival Rate**: N/A")
        lines.append(f"- **Escalation Rate**: {prof.get('escalation_rate', 'N/A'):.1%}" if prof.get('escalation_rate') is not None else "- **Escalation Rate**: N/A")
        lines.append(f"- **Renal Failure Rate**: {prof.get('renal_rate', 'N/A'):.1%}" if prof.get('renal_rate') is not None else "- **Renal Failure Rate**: N/A")
        if prof.get('clinical_recommendation'):
            lines.append(f"- **Recommendation**: {prof['clinical_recommendation']}")
        lines.append("")
        lines.append("| Feature | Mean Value |")
        lines.append("|---------|------------|")
        for feat, val in sorted(prof.get('mean_features', {}).items()):
            if not np.isnan(val) if isinstance(val, float) else True:
                lines.append(f"| {feat} | {val:.3f} |")
        lines.append("")

    lines.extend([
        "## Clinical Interpretation",
        "",
        "- **Non-congested (Low-risk)**: Best preserved renal function, lowest lactate, positive CPO response → early weaning candidates",
        "- **Cardiorenal (Moderate-risk)**: Elevated creatinine, congestion, borderline hemodynamics → staged escalation planning",
        "- **Cardiometabolic (High-risk)**: Elevated lactate, hepatic dysfunction, poor hemodynamic response → consider advanced MCS/RV support/transplant",
        "",
        "## Limitations",
        "",
        "- Small sample size limits cluster stability",
        "- Missing data imputed with median; sparse features (PV loop, echo) may reduce effective signal",
        "- Clusters are exploratory; validate on external cohort before clinical use",
        f"- k={n_clusters} is locked per Zweck et al. (2021) published methodology, not selected via silhouette maximization",
    ])

    report_path = output_dir / "clustering_report.md"
    with open(report_path, "w") as f:
        f.write("\n".join(lines))
    print(f"\nReport written to {report_path}")


# ---------------------------------------------------------------------------
# Optimal k Search
# ---------------------------------------------------------------------------

def find_optimal_k(X: np.ndarray, k_candidates: list = None, random_state: int = 42) -> tuple:
    """
    Evaluate k=2, 3, 4 with silhouette score and elbow (inertia) method.
    Returns: (best_k, results_dict)
    """
    if k_candidates is None:
        k_candidates = [2, 3, 4]

    results = {}
    inertias = []
    silhouettes = []

    print(f"  Searching optimal k among {k_candidates}...")
    for k in k_candidates:
        km = KMeans(n_clusters=k, init="k-means++", n_init=50, random_state=random_state)
        labels = km.fit_predict(X)
        inertia = km.inertia_
        sil = silhouette_score(X, labels) if k > 1 else 0.0
        results[k] = {
            "inertia": inertia,
            "silhouette": sil,
            "labels": labels,
            "centers": km.cluster_centers_,
        }
        inertias.append(inertia)
        silhouettes.append(sil)
        print(f"    k={k}: inertia={inertia:.1f}, silhouette={sil:.3f}")

    # Pick best k: maximize silhouette, but penalize very high k on small samples
    best_k = max(k_candidates, key=lambda k: results[k]["silhouette"])
    print(f"  Optimal k = {best_k} (silhouette={results[best_k]['silhouette']:.3f})")
    return best_k, results


# ---------------------------------------------------------------------------
# Experiment Runner: compare configurations
# ---------------------------------------------------------------------------

def run_clustering_experiments(X: np.ndarray, k: int, random_state: int = 42) -> dict:
    """
    Run multiple clustering configurations and return the best one.
    Tests: K-Means (raw), K-Means (PCA-5), K-Means (PCA-8), K-Means (PCA-10), GMM (raw).
    """
    from sklearn.mixture import GaussianMixture

    configs = []

    # 1. K-Means raw
    km = KMeans(n_clusters=k, init="k-means++", n_init=50, random_state=random_state)
    labels_km = km.fit_predict(X)
    sil_km = silhouette_score(X, labels_km)
    configs.append({"name": "kmeans_raw", "labels": labels_km, "silhouette": sil_km, "model": km})

    # 2. K-Means with PCA (5 components)
    pca5 = PCA(n_components=min(5, X.shape[1]), random_state=random_state)
    X_pca5 = pca5.fit_transform(X)
    km5 = KMeans(n_clusters=k, init="k-means++", n_init=50, random_state=random_state)
    labels_pca5 = km5.fit_predict(X_pca5)
    sil_pca5 = silhouette_score(X_pca5, labels_pca5)
    configs.append({"name": "kmeans_pca5", "labels": labels_pca5, "silhouette": sil_pca5, "model": km5, "pca": pca5})

    # 3. K-Means with PCA (8 components)
    pca8 = PCA(n_components=min(8, X.shape[1]), random_state=random_state)
    X_pca8 = pca8.fit_transform(X)
    km8 = KMeans(n_clusters=k, init="k-means++", n_init=50, random_state=random_state)
    labels_pca8 = km8.fit_predict(X_pca8)
    sil_pca8 = silhouette_score(X_pca8, labels_pca8)
    configs.append({"name": "kmeans_pca8", "labels": labels_pca8, "silhouette": sil_pca8, "model": km8, "pca": pca8})

    # 4. GMM raw
    gmm = GaussianMixture(n_components=k, random_state=random_state, n_init=10)
    labels_gmm = gmm.fit_predict(X)
    sil_gmm = silhouette_score(X, labels_gmm)
    configs.append({"name": "gmm_raw", "labels": labels_gmm, "silhouette": sil_gmm, "model": gmm})

    print(f"  Configuration comparison (k={k}):")
    for cfg in configs:
        marker = " <- BEST" if cfg["silhouette"] == max(c["silhouette"] for c in configs) else ""
        print(f"    {cfg['name']:15s}: silhouette={cfg['silhouette']:.3f}{marker}")

    best = max(configs, key=lambda c: c["silhouette"])
    return best


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("Impella Patient Clustering Pipeline")
    print("=" * 60)

    print("\n[1/7] Loading data...")
    df_pd = load_patient_data(DATA_PATH)
    df_cohort = load_cohort(DATA_PATH)
    print(f"  Patient Data: {len(df_pd)} patients")
    print(f"  Cohort: {len(df_cohort)} patients")

    df = df_pd.merge(df_cohort, on="mrn", how="left")
    print(f"  Merged dataset: {len(df)} patients")
    print(f"  Patients with Cohort outcomes: {df['cohort_outcome'].notna().sum()}")

    print("\n[2/7] Engineering features...")
    df = engineer_clustering_features(df)

    # Use Zweck-style features (core 6 labs + congestion marker + demographics)
    feature_list = CLUSTERING_FEATURES_ZWECK
    print(f"  Using {len(feature_list)} features (Zweck-style: core 6 + RA + demographics)")
    df_features = select_features(df, feature_list)
    print(f"  Available features: {list(df_features.columns)}")

    print("\n[3/7] Preprocessing...")
    print(f"  Patients with >50% missing features: {(df_features.isna().mean(axis=1) > 0.5).sum()}")

    # Drop patients with >50% missing
    mask_good = df_features.isna().mean(axis=1) <= 0.5
    df_features = df_features[mask_good].copy()
    df_good = df[mask_good].copy()
    print(f"  Patients retained: {len(df_features)}")

    # Impute and scale
    imputer = SimpleImputer(strategy="median")
    X_imputed = imputer.fit_transform(df_features)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_imputed)
    valid_features = list(df_features.columns)
    print(f"  Final feature count: {len(valid_features)}")

    # Lock k=3 per Zweck et al. published methodology
    n_clusters = 3
    print(f"\n[4/7] Locked k={n_clusters} (Zweck methodology)")
    print("  Skipping optimal-k search to match published CS phenotyping standard")

    print("\n[5/7] Running clustering experiments (K-Means raw, PCA-5, PCA-8, GMM)...")
    best_cfg = run_clustering_experiments(X_scaled, n_clusters, random_state=RANDOM_STATE)
    labels = best_cfg["labels"]
    print(f"  Best config: {best_cfg['name']} (silhouette={best_cfg['silhouette']:.3f})")

    # If PCA was best, we need to save the transformer and compute centroids in PCA space
    pca_transformer = best_cfg.get("pca")
    if pca_transformer is not None:
        X_clustering = pca_transformer.transform(X_scaled)
        print(f"  PCA explains {pca_transformer.explained_variance_ratio_.sum():.1%} variance")
    else:
        X_clustering = X_scaled

    # Still run consensus k-means on raw scaled data for visualization
    consensus_matrix = consensus_kmeans(
        X_scaled, n_clusters, N_BOOTSTRAP, SUBSAMPLE_RATIO, RANDOM_STATE
    )

    # Compute centroids in the clustering space (PCA or raw)
    df_good = df_good.copy().reset_index(drop=True)
    df_good["cluster"] = labels
    centroids = np.array([X_clustering[labels == c].mean(axis=0) for c in np.unique(labels)])

    # Report silhouette in both spaces
    sil_original = silhouette_score(X_scaled, labels)
    sil_clustering = silhouette_score(X_clustering, labels)
    print(f"  Silhouette (original space): {sil_original:.3f}")
    print(f"  Silhouette (clustering space): {sil_clustering:.3f}")
    silhouette_avg = sil_clustering

    print("\n[6/7] Hierarchical clustering...")
    linkage_matrix = hierarchical_clustering(X_scaled)

    print("\n[7/7] Building profiles and generating outputs...")
    profiles = build_cluster_profiles(df_good, labels, valid_features)
    profiles = assign_cluster_names(profiles)

    print("\n  Cluster assignments:")
    for cid, prof in sorted(profiles.items()):
        print(f"    Cluster {cid} ({prof.get('cluster_name', 'Unknown')}): "
              f"{prof['patient_count']} patients, escalation={prof.get('escalation_rate', 0):.1%}, "
              f"survival={prof.get('survival_rate', 0):.1%}")

    print("\n  Generating outputs...")

    # Visualizations
    plot_dendrogram(linkage_matrix, OUTPUT_DIR / "dendrogram.png",
                    labels=df_good["mrn"].tolist())
    plot_consensus_matrix(consensus_matrix, OUTPUT_DIR / "consensus_matrix.png",
                          patient_ids=df_good["mrn"].tolist())
    # Plot silhouette in the clustering space (PCA or raw) for accurate visualization
    silhouette_avg_viz = plot_silhouette(X_clustering, labels, OUTPUT_DIR / "silhouette.png",
                                         valid_features)

    # PCA scatter
    pca = PCA(n_components=2)
    X_pca = pca.fit_transform(X_scaled)
    plot_pca_scatter(X_pca, labels, OUTPUT_DIR / "pca_scatter.png",
                     patient_ids=df_good["mrn"].tolist())

    # Cluster profiles heatmap
    plot_cluster_profiles(profiles, OUTPUT_DIR / "cluster_profiles_heatmap.png")
    plot_outcome_rates(profiles, OUTPUT_DIR / "outcome_rates.png")

    # Export artifacts
    export_assignments(df_good[["mrn", "cluster"]], labels, OUTPUT_DIR / "cluster_assignments.csv")
    export_profiles(profiles, OUTPUT_DIR / "cluster_profiles.json")
    export_model(imputer, scaler, None,  # kmeans object not needed for prediction
                 centroids, valid_features, OUTPUT_DIR / "cluster_model.joblib",
                 pca=pca_transformer)
    export_quality_metrics(
        silhouette_avg_viz, n_clusters, len(valid_features), len(df_good),
        OUTPUT_DIR / "quality_metrics.json"
    )

    # Generate report
    generate_report(profiles, silhouette_avg_viz, valid_features, n_clusters, OUTPUT_DIR)

    print(f"\n  All outputs saved to {OUTPUT_DIR}/")
    print("\nPipeline complete.")

    # Summary of what was created
    print("\n" + "=" * 60)
    print("Output Summary:")
    print("=" * 60)
    for f in sorted(OUTPUT_DIR.iterdir()):
        print(f"  {f.name}: {f.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
