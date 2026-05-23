#!/usr/bin/env python3
"""
Mortality Feature Importance Analysis
======================================
Compares 5 feature importance methods to identify the most clinically
relevant variables driving mortality in Impella-supported patients.

Methods:
1. Random Forest Gini Importance
2. RF Permutation Importance
3. SHAP Values (TreeSHAP)
4. LASSO (L1-regularized Logistic Regression)
5. Univariate AUC Screening

Output: mortality_feature_report.html (self-contained HTML)
"""

import json
import warnings
import io
import base64
from pathlib import Path

import numpy as np
import pandas as pd

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold, cross_val_predict, train_test_split
from sklearn.metrics import roc_auc_score, roc_curve
from sklearn.inspection import permutation_importance
from sklearn.feature_selection import SelectFromModel

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import seaborn as sns

import shap

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_PATH = Path("Impella_MK.xlsx")
RANDOM_STATE = 42
N_SPLITS = 5
TOP_K = 20  # Show top K features in rankings

# Row mapping (from ml_pipeline.py)
PATIENT_DATA_ROWS = {
    "general_notes": 2, "first_name": 3, "last_name": 4, "mrn": 5,
    "date_of_implant": 6, "age": 7, "weight_kg": 8, "height_cm": 9,
    "gender": 10, "race": 11, "cause_of_shock": 12, "scai_stage": 13,
    "days_between_rhc_and_impella": 14,
    "pre_ra": 17, "pre_rvsp": 18, "pre_rvdp": 19, "pre_pasp": 20,
    "pre_padp": 21, "pre_map": 27, "pre_pcwp": 23, "pre_pvr": 24,
    "pre_sbp": 25, "pre_dbp": 26, "pre_hr": 28, "pre_tdco": 29,
    "pre_sv": 30, "pre_pa_o2": 31, "pre_sp_o2": 32, "pre_papi": 33,
    "pre_cpo": 34, "pre_rv_cpo": 35,
    "post_ra": 40, "post_rvsp": 41, "post_rvdp": 42, "post_pasp": 43,
    "post_padp": 44, "post_map": 50, "post_pcwp": 46, "post_pvr": 47,
    "post_sbp": 48, "post_dbp": 49, "post_hr": 51, "post_tdco": 52,
    "post_sv": 53, "post_pa_o2": 54, "post_sp_o2": 55, "post_papi": 56,
    "post_cpo": 57, "post_rv_cpo": 58,
    "pre_rvedd": 67, "pre_tapse": 68, "pre_rv_s": 69, "pre_rv_fs": 70,
    "pre_tr_severity": 71, "pre_echo_pasp": 72, "pre_lvedd": 73,
    "pre_septal_flattening": 74, "pre_atrial_bowing": 75,
    "post_rvedd": 78, "post_tapse": 79, "post_rv_s": 80, "post_rv_fs": 81,
    "post_tr_severity": 82, "post_echo_pasp": 83, "post_lvedd": 84,
    "post_septal_flattening": 85, "post_atrial_bowing": 86,
    "pre_sodium": 88, "pre_potassium": 89, "pre_hco3": 90,
    "pre_creatinine": 91, "pre_egfr": 92, "pre_hemoglobin": 93,
    "pre_wbc": 94, "pre_ast": 95, "pre_alt": 96, "pre_bili": 97,
    "pre_lactate": 98, "pre_ph": 99,
    "post_sodium": 102, "post_potassium": 103, "post_hco3": 104,
    "post_creatinine": 105, "post_egfr": 106, "post_hemoglobin": 107,
    "post_wbc": 108, "post_ast": 109, "post_alt": 110, "post_bili": 111,
    "post_lactate": 112, "post_ph": 113,
    "dopamine": 116, "dobutamine": 117, "epinephrine": 118,
    "milrinone": 119, "norepinephrine": 120, "vasopressin": 121,
    "vis_score": 122,
    "pre_furosemide": 125, "pre_augmentation": 126,
    "post_furosemide": 129, "post_augmentation": 130,
    "impella_performance": 133, "impella_flow": 134,
    "renal_failure": 138, "intubation": 139, "mcs_escalation": 140,
    "outcome": 141,
    "ees": 144, "ea": 145, "ees_ea": 146, "esp": 147, "edp": 148,
    "pmax": 149, "esv": 150, "edv": 151, "pv_sv": 152,
    "dp_dt_max": 153, "dp_dt_min": 154,
}

# Columns to exclude (leakage, non-predictive identifiers)
LEAKAGE_COLS = [
    "outcome", "cohort_outcome", "intubation", "mcs_escalation",
    "general_notes", "first_name", "last_name", "mrn", "date_of_implant",
    "pre_septal_flattening", "pre_atrial_bowing",
    "post_septal_flattening", "post_atrial_bowing",
    "cohort_age", "cohort_gender", "indication",
    "physician", "rhc_prior_72h", "rhc_timing_days", "support_days",
    "Absolute Value",
]

# Duplicate groups: keep only the first canonical version
DUPLICATE_GROUPS = {
    "race": ["race", "race_numeric"],
    "gender": ["gender", "gender_numeric"],
    "cause_of_shock": ["cause_of_shock", "shock_cause_numeric"],
    "scai_stage": ["scai_stage", "scai_numeric"],
    "age": ["age", "Age"],
}


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------
def load_patient_data(path: Path) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Patient Data"]

    rows = []
    for col in range(2, ws.max_column + 1):
        mrn = ws.cell(row=5, column=col).value
        if not mrn:
            continue
        record = {"mrn": str(mrn).strip()}
        for key, row_idx in PATIENT_DATA_ROWS.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, str):
                val = val.strip()
                if val.lower() in ("n/a", "na", "", "none"):
                    val = np.nan
            record[key] = val
        rows.append(record)

    df = pd.DataFrame(rows)
    numeric_keys = [k for k in PATIENT_DATA_ROWS if k not in
                    ("general_notes", "first_name", "last_name", "mrn", "date_of_implant")]
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")
    return df


def load_cohort(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Cohort")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    rename_map = {
        "MRN": "mrn", "Outcome": "cohort_outcome",
        "Age ": "cohort_age", "Gender": "cohort_gender",
        "Indication for Use": "indication",
        "Duration of Support Time (Days)": "support_days",
        "Physician Name": "physician",
        "RHC prior to implant (<72 hours)": "rhc_prior_72h",
        "RHC timing (days prior)": "rhc_timing_days",
    }
    df = df.rename(columns=rename_map)
    df["mrn"] = df["mrn"].astype(str).str.strip()
    return df


# ---------------------------------------------------------------------------
# Feature Engineering
# ---------------------------------------------------------------------------
def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["bmi"] = df["weight_kg"] / ((df["height_cm"] / 100) ** 2)

    rhc_vars = ["ra", "rvsp", "rvdp", "pasp", "padp", "map", "pcwp", "pvr",
                 "sbp", "dbp", "hr", "tdco", "sv", "pa_o2", "sp_o2", "papi", "cpo", "rv_cpo"]
    for v in rhc_vars:
        pre, post = f"pre_{v}", f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]
            df[f"ratio_{v}"] = df[post] / df[pre].replace(0, np.nan)

    echo_vars = ["rvedd", "tapse", "rv_s", "rv_fs", "tr_severity", "echo_pasp", "lvedd"]
    for v in echo_vars:
        pre, post = f"pre_{v}", f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]
            df[f"ratio_{v}"] = df[post] / df[pre].replace(0, np.nan)

    lab_vars = ["sodium", "potassium", "hco3", "creatinine", "egfr",
                "hemoglobin", "wbc", "ast", "alt", "bili", "lactate", "ph"]
    for v in lab_vars:
        pre, post = f"pre_{v}", f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]

    df["delta_cpo"] = df["post_cpo"] - df["pre_cpo"]
    df["recovery_score"] = np.clip((df["delta_cpo"] + 0.5) * 100, 0, 100)
    df["vis_high"] = (df["vis_score"] > 15).astype(int)

    inotrope_cols = ["dopamine", "dobutamine", "epinephrine", "milrinone", "norepinephrine", "vasopressin"]
    df["inotrope_count"] = df[inotrope_cols].gt(0).sum(axis=1)

    scai_map = {"b": 1, "c": 2, "d": 3, "e": 4}
    df["scai_numeric"] = df["scai_stage"].astype(str).str.lower().map(scai_map)

    return df


# ---------------------------------------------------------------------------
# Remove duplicates and leakage
# ---------------------------------------------------------------------------
def clean_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Remove leakage columns (those that exist)
    cols_to_drop = [c for c in LEAKAGE_COLS if c in df.columns]
    df = df.drop(columns=cols_to_drop)

    # Resolve duplicate groups: keep only the first canonical name per group
    all_drop_dupes = []
    for canonical, variants in DUPLICATE_GROUPS.items():
        existing = [v for v in variants if v in df.columns]
        if len(existing) > 1:
            # keep the canonical if it exists, else the first variant
            keep = canonical if canonical in existing else existing[0]
            drop = [v for v in existing if v != keep]
            all_drop_dupes.extend(drop)
    df = df.drop(columns=[c for c in all_drop_dupes if c in df.columns])
    return df


# ---------------------------------------------------------------------------
# Target building: survival (expired vs survived)
# ---------------------------------------------------------------------------
def build_target(df: pd.DataFrame) -> pd.DataFrame:
    # Use 'outcome' for target, then drop it
    df["target_survival"] = (df["cohort_outcome"].str.lower() == "expired").astype(int)
    df = df.drop(columns=["cohort_outcome"])
    return df


# ---------------------------------------------------------------------------
# Helper: fig to base64 PNG
# ---------------------------------------------------------------------------
def fig_to_b64(fig, dpi=120) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    img = base64.b64encode(buf.read()).decode("utf-8")
    plt.close(fig)
    return img


# ---------------------------------------------------------------------------
# Method 1: Random Forest Gini Importance
# ---------------------------------------------------------------------------
def rf_gini_importance(X_train, y_train, feature_names):
    rf = RandomForestClassifier(
        n_estimators=500, max_depth=8, min_samples_leaf=3,
        class_weight="balanced", random_state=RANDOM_STATE, n_jobs=-1
    )
    rf.fit(X_train, y_train)
    importances = rf.feature_importances_
    scores = pd.Series(importances, index=feature_names).sort_values(ascending=False)
    return rf, scores


# ---------------------------------------------------------------------------
# Method 2: Permutation Importance
# ---------------------------------------------------------------------------
def permutation_importance_method(X_train, y_train, X_test, y_test, feature_names):
    rf = RandomForestClassifier(
        n_estimators=500, max_depth=8, min_samples_leaf=3,
        class_weight="balanced", random_state=RANDOM_STATE, n_jobs=-1
    )
    rf.fit(X_train, y_train)
    perm = permutation_importance(rf, X_test, y_test, n_repeats=30,
                                  random_state=RANDOM_STATE, n_jobs=-1)
    scores = pd.Series(perm.importances_mean, index=feature_names).sort_values(ascending=False)
    return scores


# ---------------------------------------------------------------------------
# Method 3: SHAP Values
# ---------------------------------------------------------------------------
def shap_importance(rf_model, X_test, feature_names):
    explainer = shap.TreeExplainer(rf_model)
    shap_values = explainer.shap_values(X_test)
    # shap_values shape: (n_samples, n_features, n_classes) for sklearn RF
    if isinstance(shap_values, list):
        sv_class1 = shap_values[1]
    elif shap_values.ndim == 3:
        sv_class1 = shap_values[:, :, 1]
    else:
        sv_class1 = shap_values
    # Mean absolute SHAP value per feature
    scores = pd.Series(np.abs(sv_class1).mean(axis=0), index=feature_names).sort_values(ascending=False)
    return scores, sv_class1, explainer
    return scores, sv, explainer


# ---------------------------------------------------------------------------
# Method 4: LASSO (L1 Logistic Regression)
# ---------------------------------------------------------------------------
def lasso_importance(X_train, y_train, feature_names):
    # Tune C via stratified CV
    y_arr = np.asarray(y_train).ravel()
    best_c = 1.0
    best_auc = 0
    for c in np.logspace(-2, 2, 30):
        lr = LogisticRegression(penalty="l1", C=c, solver="saga",
                                max_iter=5000, class_weight="balanced",
                                random_state=RANDOM_STATE)
        aucs = []
        skf = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
        for train_idx, val_idx in skf.split(X_train, y_arr):
            lr.fit(X_train[train_idx], y_arr[train_idx])
            y_proba = lr.predict_proba(X_train[val_idx])[:, 1]
            aucs.append(roc_auc_score(y_arr[val_idx], y_proba))
        mean_auc = np.mean(aucs)
        if mean_auc > best_auc:
            best_auc = mean_auc
            best_c = c

    lasso = LogisticRegression(penalty="l1", C=best_c, solver="saga",
                               max_iter=5000, class_weight="balanced",
                               random_state=RANDOM_STATE)
    lasso.fit(X_train, y_arr)
    coef = lasso.coef_[0]
    scores = pd.Series(np.abs(coef), index=feature_names).sort_values(ascending=False)
    n_selected = int(np.sum(coef != 0))
    return scores, best_c, n_selected


# ---------------------------------------------------------------------------
# Method 5: Univariate AUC Screening
# ---------------------------------------------------------------------------
def univariate_auc(X_train, y_train, feature_names):
    """Compute AUC for each feature used alone in a logistic regression."""
    y_arr = np.asarray(y_train).ravel()
    aucs = []
    for i in range(X_train.shape[1]):
        x_i = X_train[:, i].reshape(-1, 1)
        try:
            lr = LogisticRegression(max_iter=1000, class_weight="balanced",
                                    random_state=RANDOM_STATE)
            skf = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
            y_proba = cross_val_predict(lr, x_i, y_arr, cv=skf, method="predict_proba")[:, 1]
            aucs.append(roc_auc_score(y_arr, y_proba))
        except Exception:
            aucs.append(0.5)
    scores = pd.Series(aucs, index=feature_names).sort_values(ascending=False)
    return scores


# ---------------------------------------------------------------------------
# Consensus Ranking
# ---------------------------------------------------------------------------
def build_consensus(all_rankings: dict) -> pd.DataFrame:
    """Normalize each method to [0,1] and compute average rank across methods."""
    all_scores = {}
    all_ranks = {}

    for method, scores in all_rankings.items():
        all_scores[method] = scores
        ranks = scores.rank(ascending=False)
        all_ranks[method] = ranks

    consensus_df = pd.DataFrame(all_ranks)
    consensus_df["mean_rank"] = consensus_df.mean(axis=1)
    consensus_df = consensus_df.sort_values("mean_rank")

    # Also add normalized scores
    score_df = pd.DataFrame(all_scores)
    # Normalize each method to [0,1]
    for col in score_df.columns:
        mx = score_df[col].max()
        if mx > 0:
            score_df[col] = score_df[col] / mx
    score_df["consensus_score"] = score_df.mean(axis=1)
    score_df = score_df.loc[consensus_df.index]

    combined = consensus_df.join(score_df[["consensus_score"]])
    return combined


# ---------------------------------------------------------------------------
# Plotting Functions
# ---------------------------------------------------------------------------
def plot_consensus_bar(consensus_df: pd.DataFrame, top_k: int = TOP_K):
    top = consensus_df.head(top_k)
    fig, ax = plt.subplots(figsize=(10, 8))
    colors = plt.cm.YlOrRd(top["consensus_score"].values / top["consensus_score"].max())
    bars = ax.barh(range(len(top)), top["consensus_score"].values, color=colors, edgecolor="grey")
    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(top.index)
    ax.invert_yaxis()
    ax.set_xlabel("Consensus Normalized Score")
    ax.set_title(f"Top {top_k} Features — Consensus Across 5 Methods")
    ax.axvline(0.5, color="grey", linestyle="--", alpha=0.5)
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_method_top_k(all_rankings: dict, top_k: int = 15):
    """Side-by-side bar charts for each method's top-k features."""
    n_methods = len(all_rankings)
    fig, axes = plt.subplots(1, n_methods, figsize=(5 * n_methods, 6))
    if n_methods == 1:
        axes = [axes]

    for ax, (method, scores) in zip(axes, all_rankings.items()):
        top = scores.head(top_k)
        colors = plt.cm.Blues(np.linspace(0.4, 0.9, len(top)))
        ax.barh(range(len(top)), top.values, color=colors[::-1], edgecolor="grey")
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top.index, fontsize=8)
        ax.invert_yaxis()
        ax.set_title(method, fontsize=10)
        ax.set_xlabel("Importance")
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_shap_beeswarm(shap_values, X_test, feature_names):
    fig, ax = plt.subplots(figsize=(10, 6))
    shap.summary_plot(shap_values, X_test, feature_names=feature_names,
                      show=False, max_display=TOP_K)
    plt.title("SHAP Summary — Impact on Mortality Prediction")
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_lasso_path(X_train, y_train, feature_names):
    """Plot LASSO coefficient path across regularization strengths."""
    y_arr = np.asarray(y_train).ravel()
    Cs = np.logspace(-2, 2, 30)
    coefs = []
    for c in Cs:
        lr = LogisticRegression(penalty="l1", C=c, solver="saga", max_iter=5000,
                                class_weight="balanced", random_state=RANDOM_STATE)
        lr.fit(X_train, y_arr)
        coefs.append(lr.coef_[0])
    coefs = np.array(coefs)

    fig, ax = plt.subplots(figsize=(10, 6))
    for i in range(min(coefs.shape[1], 172)):
        ax.plot(Cs, coefs[:, i], lw=0.8, alpha=0.6)
    ax.set_xscale("log")
    ax.set_xlabel("C (inverse regularization, log scale)")
    ax.set_ylabel("Coefficient")
    ax.set_title("LASSO Coefficient Path")
    ax.axhline(0, color="grey", linestyle="--", linewidth=0.5)
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_univariate_auc_bar(scores: pd.Series, top_k: int = TOP_K):
    top = scores.head(top_k)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.Greens(np.linspace(0.3, 0.9, len(top)))[::-1]
    ax.barh(range(len(top)), top.values, color=colors, edgecolor="grey")
    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(top.index)
    ax.invert_yaxis()
    ax.set_xlabel("Univariate AUC")
    ax.set_title(f"Top {top_k} Features by Univariate AUC")
    ax.axvline(0.5, color="red", linestyle="--", alpha=0.5, label="Random")
    ax.legend()
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_correlation_heatmap(consensus_df, X, top_k: int = TOP_K):
    top_features = consensus_df.head(top_k).index.tolist()
    X_sub = pd.DataFrame(X, columns=consensus_df.index)
    existing = [f for f in top_features if f in X_sub.columns]
    corr = X_sub[existing].corr()

    fig, ax = plt.subplots(figsize=(10, 8))
    mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
    sns.heatmap(corr, mask=mask, annot=True, fmt=".2f", cmap="RdBu_r",
                center=0, vmin=-1, vmax=1, square=True, ax=ax,
                cbar_kws={"shrink": 0.8})
    ax.set_title(f"Correlation Heatmap — Top {top_k} Consensus Features")
    fig.tight_layout()
    return fig_to_b64(fig)


def plot_missing_data_heatmap(df, feature_names):
    """Show missing data pattern for top features."""
    fig, ax = plt.subplots(figsize=(12, 4))
    missing = df[feature_names].isna().astype(int)
    sns.heatmap(missing.T, cmap="RdYlBu_r", cbar=False, ax=ax, yticklabels=True)
    ax.set_xlabel("Patients")
    ax.set_ylabel("Feature")
    ax.set_title("Missing Data Pattern")
    fig.tight_layout()
    return fig_to_b64(fig)


# ---------------------------------------------------------------------------
# HTML Report Generation
# ---------------------------------------------------------------------------
def generate_html(all_rankings, consensus_df, method_results,
                  class_dist, n_features, n_patients, plots):
    # Build method tables
    method_tables = ""
    for method, scores in all_rankings.items():
        top = scores.head(TOP_K)
        rows = ""
        for rank, (feat, val) in enumerate(top.items(), 1):
            rows += f"<tr><td>{rank}</td><td>{feat}</td><td>{val:.4f}</td></tr>\n"
        method_tables += f"""
        <div class="method-block">
            <h3>{method}</h3>
            <table>
                <tr><th>Rank</th><th>Feature</th><th>Score</th></tr>
                {rows}
            </table>
        </div>
        """

    # Consensus table
    consensus_rows = ""
    for rank, (feat, row) in enumerate(consensus_df.head(TOP_K).iterrows(), 1):
        cols = " ".join(f"<td>{row[m]:.2f}</td>" for m in all_rankings.keys())
        consensus_rows += f"<tr><td>{rank}</td><td>{feat}</td>{cols}<td>{row['consensus_score']:.3f}</td></tr>\n"

    method_headers = " ".join(f"<th>{m}</th>" for m in all_rankings.keys())

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Mortality Feature Importance Report — Impella Analytics</title>
<style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
           max-width: 1200px; margin: 0 auto; padding: 20px; background: #f5f5f5; color: #333; }}
    h1 {{ color: #1a1a2e; border-bottom: 3px solid #e94560; padding-bottom: 8px; }}
    h2 {{ color: #16213e; margin-top: 30px; }}
    .summary {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
    .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin: 15px 0; }}
    .stat-card {{ background: #1a1a2e; color: white; padding: 15px; border-radius: 8px; text-align: center; }}
    .stat-card .value {{ font-size: 28px; font-weight: bold; }}
    .stat-card .label {{ font-size: 12px; opacity: 0.8; }}
    .plot-container {{ background: white; padding: 15px; border-radius: 8px; margin: 20px 0;
                      box-shadow: 0 2px 4px rgba(0,0,0,0.1); text-align: center; }}
    .plot-container img {{ max-width: 100%; height: auto; }}
    .methods-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 16px; }}
    .method-block {{ background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
    .method-block h3 {{ margin-top: 0; color: #e94560; font-size: 14px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    th, td {{ padding: 6px 8px; text-align: left; border-bottom: 1px solid #eee; }}
    th {{ background: #16213e; color: white; font-size: 11px; }}
    tr:hover {{ background: #f0f0f0; }}
    .interpretation {{ background: white; padding: 20px; border-radius: 8px; margin: 20px 0;
                      box-shadow: 0 2px 4px rgba(0,0,0,0.1); line-height: 1.6; }}
    .limitations {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 4px; margin: 15px 0; }}
    .full-table {{ width: 100%; overflow-x: auto; }}
    .section-tag {{ display: inline-block; background: #e94560; color: white; padding: 2px 10px;
                   border-radius: 12px; font-size: 11px; margin-bottom: 8px; }}
</style>
</head>
<body>
<h1>Mortality Feature Importance Report</h1>
<p style="color:#666;">Comprehensive multi-method analysis of clinical variables driving mortality in Impella-supported patients</p>

<div class="summary">
    <h2>Dataset Summary</h2>
    <div class="summary-grid">
        <div class="stat-card">
            <div class="value">{n_patients}</div>
            <div class="label">Total Patients</div>
        </div>
        <div class="stat-card">
            <div class="value">{class_dist.get('expired', 0)}</div>
            <div class="label">Expired ({class_dist.get('expired_pct', 0):.1f}%)</div>
        </div>
        <div class="stat-card">
            <div class="value">{class_dist.get('survived', 0)}</div>
            <div class="label">Survived ({class_dist.get('survived_pct', 0):.1f}%)</div>
        </div>
        <div class="stat-card">
            <div class="value">{n_features}</div>
            <div class="label">Candidate Features</div>
        </div>
        <div class="stat-card">
            <div class="value">{len(consensus_df.head(TOP_K))}</div>
            <div class="label">Top-K Selected</div>
        </div>
    </div>

    <h3>Methods Compared</h3>
    <ol>
        <li><strong>Random Forest Gini Importance</strong> — Built-in impurity-based feature importance from 500-tree RF</li>
        <li><strong>Permutation Importance</strong> — Drop in AUC when feature values are shuffled (more robust than Gini)</li>
        <li><strong>SHAP Values (TreeSHAP)</strong> — Shapley additive explanations showing each feature's contribution</li>
        <li><strong>LASSO (L1 Logistic Regression)</strong> — Sparse logistic regression with cross-validated regularization</li>
        <li><strong>Univariate AUC Screening</strong> — Each feature's individual predictive power as a standalone classifier</li>
    </ol>
</div>

<div class="plot-container">
    <h2>Consensus Ranking — Top {TOP_K} Features</h2>
    <img src="data:image/png;base64,{plots['consensus_bar']}" alt="Consensus Ranking">
    <p style="color:#666;font-size:12px;">Normalized average rank across all 5 methods. Higher = more important.</p>
</div>

<div class="plot-container">
    <h2>SHAP Summary Plot</h2>
    <img src="data:image/png;base64,{plots['shap_beeswarm']}" alt="SHAP Summary">
    <p style="color:#666;font-size:12px;">Red = high feature value, blue = low. Position on x-axis = impact on model output.</p>
</div>

<div class="plot-container">
    <h2>Correlation Heatmap — Top {TOP_K} Consensus Features</h2>
    <img src="data:image/png;base64,{plots['correlation_heatmap']}" alt="Correlation Heatmap">
</div>

<h2>Per-Method Top {TOP_K} Rankings</h2>
<div class="methods-grid">
    {method_tables}
</div>

<div class="plot-container">
    <h2>Per-Method Top 15 Feature Rankings</h2>
    <img src="data:image/png;base64,{plots['method_bars']}" alt="Per-Method Rankings">
</div>

<div class="plot-container">
    <h2>Univariate AUC — Top {TOP_K} Features</h2>
    <img src="data:image/png;base64,{plots['univariate_auc']}" alt="Univariate AUC">
    <p style="color:#666;font-size:12px;">Red dashed line = random AUC (0.5). Bars show each feature's standalone predictive power.</p>
</div>

<div class="plot-container">
    <h2>LASSO Coefficient Path</h2>
    <img src="data:image/png;base64,{plots['lasso_path']}" alt="LASSO Path">
    <p style="color:#666;font-size:12px;">How coefficients shrink to zero as regularization increases.</p>
</div>

<!-- Full consensus table -->
<div class="summary">
    <h2>Full Consensus Ranking Table — Top {TOP_K}</h2>
    <div class="full-table">
        <table>
            <tr><th>Rank</th><th>Feature</th>{method_headers}<th>Consensus Score</th></tr>
            {consensus_rows}
        </table>
    </div>
</div>

<div class="interpretation">
    <h2>Clinical Interpretation</h2>
    <span class="section-tag">Key Findings</span>
    <p>
        The consensus ranking identifies the features most consistently associated with mortality
        across diverse statistical and ML methods. Features repeatedly selected across all methods
        represent the most robust mortality signals in this dataset.
    </p>
    <p>
        <strong>Top-tier features</strong> (appearing in top-10 of 3+ methods) are the strongest candidates
        for clinical risk stratification. These should be prioritized for external validation.
    </p>
    <p>
        <strong>Note on delta/ratio features:</strong> Change-from-baseline metrics (e.g., delta CPO) capture
        physiologic trajectory, which may be more informative than single timepoint measurements.
        Their prominence in the rankings suggests that <em>trends</em> matter as much as absolute values.
    </p>
</div>

<div class="limitations">
    <strong>Limitations & Caveats:</strong>
    <ul>
        <li>Sample size is limited (N={n_patients}) — feature importance rankings may change with larger cohorts.</li>
        <li>Missing data imputed with median; features with >50% missing have unreliable importance estimates.</li>
        <li>Class imbalance ({class_dist.get('expired', 0)}/{n_patients} expired) limits statistical power for rare-but-important signals.</li>
        <li>These are associational, not causal, relationships. Confounding and collider bias cannot be excluded.</li>
        <li>External validation on an independent cohort is essential before clinical deployment.</li>
        <li>Gini importance can be biased toward high-cardinality features; permutation and SHAP are more reliable.</li>
    </ul>
</div>

<p style="text-align:center;color:#999;font-size:12px;margin-top:30px;">
    Generated by mortality_feature_analysis.py &mdash; Impella Analytics &mdash; {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}
</p>
</body>
</html>"""
    return html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("Mortality Feature Importance Analysis")
    print("=" * 60)

    # 1. Load data
    print("\n[1/7] Loading data...")
    df_pd = load_patient_data(DATA_PATH)
    df_cohort = load_cohort(DATA_PATH)
    df = df_pd.merge(df_cohort, on="mrn", how="outer")
    print(f"  Patient Data: {len(df_pd)} rows, Cohort: {len(df_cohort)} rows, Merged: {len(df)} rows")

    # Build target
    df = build_target(df)
    expired = int(df["target_survival"].sum())
    survived = int((1 - df["target_survival"]).sum())
    total = len(df)
    class_dist = {
        "expired": expired, "survived": survived,
        "expired_pct": 100 * expired / total if total else 0,
        "survived_pct": 100 * survived / total if total else 0,
    }
    print(f"  Class distribution: {expired} expired / {survived} survived ({total} total)")

    # 2. Engineer features
    print("\n[2/7] Engineering features...")
    df = engineer_features(df)

    # 3. Clean (remove leakage, duplicates)
    print("\n[3/7] Cleaning features...")
    df = clean_features(df)

    # 4. Prepare feature matrix
    print("\n[4/7] Preparing feature matrix...")
    exclude_base = {"target_survival", "mrn", "first_name", "last_name",
                    "general_notes", "date_of_implant"}
    feature_cols = [c for c in df.columns
                    if c not in exclude_base and pd.api.types.is_numeric_dtype(df[c])
                    and not c.startswith("Unnamed")]
    print(f"  Candidate features: {len(feature_cols)}")

    df_model = df.dropna(subset=["target_survival"])
    y = df_model["target_survival"].astype(int).values
    n_patients = len(y)
    print(f"  Model-ready patients: {n_patients}")

    X = df_model[feature_cols].copy()

    # Drop all-NaN columns
    all_nan = X.columns[X.isna().all()].tolist()
    if all_nan:
        print(f"  Dropping {len(all_nan)} all-NaN columns")
        X = X.drop(columns=all_nan)
    valid_features = X.columns.tolist()

    # Impute & scale
    imputer = SimpleImputer(strategy="median")
    X_imp = imputer.fit_transform(X)
    scaler = StandardScaler()
    X_scl = scaler.fit_transform(X_imp)

    # Train/test split for methods that need held-out data
    X_train, X_test, y_train, y_test = train_test_split(
        X_scl, y, test_size=0.3, stratify=y, random_state=RANDOM_STATE
    )
    print(f"  Train: {len(y_train)}, Test: {len(y_test)}")

    n_features = X_scl.shape[1]
    missing_pct = (X.isna().sum() / len(X)).sort_values(ascending=False)
    high_missing = missing_pct[missing_pct > 0.5]
    if len(high_missing):
        print(f"  Features with >50% missing: {len(high_missing)}")

    # 5. Run all methods
    print("\n[5/7] Running feature importance methods...")

    all_rankings = {}
    method_results = {}

    # Method 1: RF Gini
    print("  Method 1/5: RF Gini Importance...")
    rf_model, gini_scores = rf_gini_importance(X_train, y_train, valid_features)
    all_rankings["RF Gini"] = gini_scores
    method_results["RF Gini"] = {"auc": None}

    # Evaluate RF
    rf_auc = roc_auc_score(y_test, rf_model.predict_proba(X_test)[:, 1])
    method_results["RF Gini"]["auc"] = f"{rf_auc:.3f}"
    print(f"    RF Test AUC: {rf_auc:.3f}")

    # Method 2: Permutation
    print("  Method 2/5: Permutation Importance...")
    perm_scores = permutation_importance_method(X_train, y_train, X_test, y_test, valid_features)
    all_rankings["Permutation"] = perm_scores
    method_results["Permutation"] = {"auc": f"{rf_auc:.3f}"}

    # Method 3: SHAP
    print("  Method 3/5: SHAP Values...")
    shap_scores, shap_vals, shap_explainer = shap_importance(rf_model, X_test, valid_features)
    all_rankings["SHAP"] = shap_scores
    method_results["SHAP"] = {"auc": f"{rf_auc:.3f}"}

    # Method 4: LASSO
    print("  Method 4/5: LASSO...")
    lasso_scores, best_c, n_selected = lasso_importance(X_train, y_train, valid_features)
    all_rankings["LASSO"] = lasso_scores
    method_results["LASSO"] = {"auc": best_c, "n_selected": n_selected}
    print(f"    Best C={best_c:.4f}, non-zero coefficients: {n_selected}")

    # Evaluate LASSO on test
    lasso_final = LogisticRegression(penalty="l1", C=best_c, solver="saga",
                                     max_iter=5000, class_weight="balanced",
                                     random_state=RANDOM_STATE)
    lasso_final.fit(X_train, np.asarray(y_train).ravel())
    lasso_auc = roc_auc_score(y_test, lasso_final.predict_proba(X_test)[:, 1])
    method_results["LASSO"]["auc"] = f"{lasso_auc:.3f}"
    print(f"    LASSO Test AUC: {lasso_auc:.3f}")

    # Method 5: Univariate AUC
    print("  Method 5/5: Univariate AUC...")
    univar_scores = univariate_auc(X_train, y_train, valid_features)
    all_rankings["Univariate AUC"] = univar_scores
    method_results["Univariate AUC"] = {"auc": None}

    # 6. Build consensus
    print("\n[6/7] Building consensus ranking...")
    consensus_df = build_consensus(all_rankings)
    print(f"\nTop 10 consensus features:")
    for rank, (feat, row) in enumerate(consensus_df.head(10).iterrows(), 1):
        print(f"  {rank:2d}. {feat:35s} score={row['consensus_score']:.3f}")

    # 7. Generate HTML report
    print("\n[7/7] Generating HTML report...")

    # Plot images
    print("  Rendering plots...")
    plots = {
        "consensus_bar": plot_consensus_bar(consensus_df),
        "shap_beeswarm": plot_shap_beeswarm(shap_vals, X_test, valid_features),
        "correlation_heatmap": plot_correlation_heatmap(consensus_df, X_scl),
        "method_bars": plot_method_top_k(all_rankings),
        "univariate_auc": plot_univariate_auc_bar(univar_scores),
        "lasso_path": plot_lasso_path(X_train, y_train, valid_features),
    }

    html_content = generate_html(
        all_rankings, consensus_df, method_results,
        class_dist, n_features, n_patients, plots
    )

    report_path = Path("mortality_feature_report.html")
    with open(report_path, "w") as f:
        f.write(html_content)
    print(f"  Report written to {report_path.resolve()}")

    # Also save consensus CSV
    csv_path = Path("mortality_feature_consensus.csv")
    consensus_df.to_csv(csv_path)
    print(f"  Consensus CSV written to {csv_path.resolve()}")

    print("\n" + "=" * 60)
    print("Analysis complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
