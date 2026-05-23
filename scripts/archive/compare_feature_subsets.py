#!/usr/bin/env python3
"""
Feature Subset Comparison
=========================
Tests how model performance changes when using only the top-k consensus
features vs. the full feature set for all 3 ML targets.

Reads consensus rankings from mortality_feature_consensus.csv and tests
subsets: top 5, 10, 15, 20, 30, 50, and full feature set.

Output: ml_output/feature_subset_comparison.json + PNG plots
"""

import json, warnings, io, base64
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold, cross_val_predict, cross_val_score
from sklearn.metrics import roc_auc_score

warnings.filterwarnings("ignore")

DATA_PATH = Path("Impella_MK.xlsx")
CONSENSUS_PATH = Path("mortality_feature_consensus.csv")
OUTPUT_DIR = Path("ml_output")
OUTPUT_DIR.mkdir(exist_ok=True)
RANDOM_STATE = 42
N_SPLITS = 5
SUBSETS = [5, 10, 15, 20, 30, 50]

# Row mapping (identical to mortality_feature_analysis.py)
PATIENT_DATA_ROWS = {
    "general_notes": 2, "first_name": 3, "last_name": 4, "mrn": 5,
    "date_of_implant": 6, "age": 7, "weight_kg": 8, "height_cm": 9,
    "gender": 10, "race": 11, "cause_of_shock": 12, "scai_stage": 13,
    "days_between_rhc_and_impella": 14,
    "pre_ra": 17, "pre_rvsp": 18, "pre_rvdp": 19, "pre_pasp": 20,
    "pre_padp": 21, "pre_map": 27, "pre_pcwp": 23, "pre_pvr": 24,
    "pre_sbp": 25, "pre_dbp": 26, "pre_hr": 28, "pre_tdco": 29,
    "pre_sv": 30, "pre_pa_o2": 31, "pre_sp_o2": 32, "pre_papi": 33,
    "pre_cpo": 34, "pre_rv_cpo": 35,
    "post_ra": 40, "post_rvsp": 41, "post_rvdp": 42, "post_pasp": 43,
    "post_padp": 44, "post_map": 50, "post_pcwp": 46, "post_pvr": 47,
    "post_sbp": 48, "post_dbp": 49, "post_hr": 51, "post_tdco": 52,
    "post_sv": 53, "post_pa_o2": 54, "post_sp_o2": 55, "post_papi": 56,
    "post_cpo": 57, "post_rv_cpo": 58,
    "pre_rvedd": 67, "pre_tapse": 68, "pre_rv_s": 69, "pre_rv_fs": 70,
    "pre_tr_severity": 71, "pre_echo_pasp": 72, "pre_lvedd": 73,
    "post_rvedd": 78, "post_tapse": 79, "post_rv_s": 80, "post_rv_fs": 81,
    "post_tr_severity": 82, "post_echo_pasp": 83, "post_lvedd": 84,
    "pre_sodium": 88, "pre_potassium": 89, "pre_hco3": 90,
    "pre_creatinine": 91, "pre_egfr": 92, "pre_hemoglobin": 93,
    "pre_wbc": 94, "pre_ast": 95, "pre_alt": 96, "pre_bili": 97,
    "pre_lactate": 98, "pre_ph": 99,
    "post_sodium": 102, "post_potassium": 103, "post_hco3": 104,
    "post_creatinine": 105, "post_egfr": 106, "post_hemoglobin": 107,
    "post_wbc": 108, "post_ast": 109, "post_alt": 110, "post_bili": 111,
    "post_lactate": 112, "post_ph": 113,
    "dopamine": 116, "dobutamine": 117, "epinephrine": 118,
    "milrinone": 119, "norepinephrine": 120, "vasopressin": 121,
    "vis_score": 122,
    "pre_furosemide": 125, "pre_augmentation": 126,
    "post_furosemide": 129, "post_augmentation": 130,
    "impella_performance": 133, "impella_flow": 134,
    "renal_failure": 138, "intubation": 139, "mcs_escalation": 140,
    "outcome": 141,
    "ees": 144, "ea": 145, "ees_ea": 146, "esp": 147, "edp": 148,
    "pmax": 149, "esv": 150, "edv": 151, "pv_sv": 152,
    "dp_dt_max": 153, "dp_dt_min": 154,
}


def load_and_prepare():
    import openpyxl
    wb = openpyxl.load_workbook(DATA_PATH, data_only=True)
    ws = wb["Patient Data"]
    rows = []
    for col in range(2, ws.max_column + 1):
        mrn = ws.cell(row=5, column=col).value
        if not mrn: continue
        record = {"mrn": str(mrn).strip()}
        for key, row_idx in PATIENT_DATA_ROWS.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, str):
                val = val.strip()
                if val.lower() in ("n/a", "na", "", "none"):
                    val = np.nan
            record[key] = val
        rows.append(record)
    df = pd.DataFrame(rows)
    numeric_keys = [k for k in PATIENT_DATA_ROWS if k not in
                    ("general_notes", "first_name", "last_name", "mrn", "date_of_implant")]
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")

    # Cohort
    cohort = pd.read_excel(DATA_PATH, sheet_name="Cohort")
    cohort.columns = [c.strip() if isinstance(c, str) else c for c in cohort.columns]
    cohort = cohort.rename(columns={
        "MRN": "mrn", "Outcome": "cohort_outcome",
        "Age ": "cohort_age", "Gender": "cohort_gender",
    })
    cohort["mrn"] = cohort["mrn"].astype(str).str.strip()
    df = df.merge(cohort, on="mrn", how="outer")

    # Targets
    df["target_survival"] = (df["cohort_outcome"].str.lower() == "expired").astype(int)
    notes = df["general_notes"].fillna("").str.lower()
    df["target_escalation"] = ((notes.str.contains("ecmo|lvad|transplant|arrest|rv failure|rvf", regex=True, na=False)) | (df["mcs_escalation"].fillna(0) > 0)).astype(int)
    rv_dys = pd.Series(False, index=df.index)
    rv_dys |= df["post_papi"] < 1.0
    rv_dys |= df["post_ra"] > 20
    rv_dys |= df["post_tapse"] < 1.6
    rv_dys |= df["post_rv_s"] < 9.5
    rv_dys |= notes.str.contains("rv failure|rv dysfunction|rvf", regex=True, na=False)
    df["target_rv_dysfunction"] = rv_dys.astype(int)

    # Feature engineering
    df["bmi"] = df["weight_kg"] / ((df["height_cm"] / 100) ** 2)
    for prefix, vars_list in [("", ["ra", "rvsp", "rvdp", "pasp", "padp", "map", "pcwp", "pvr",
                                     "sbp", "dbp", "hr", "tdco", "sv", "pa_o2", "sp_o2", "papi", "cpo", "rv_cpo"]),
                               ("", ["rvedd", "tapse", "rv_s", "rv_fs", "tr_severity", "echo_pasp", "lvedd"])]:
        for v in vars_list:
            pre, post = f"pre_{v}", f"post_{v}"
            if pre in df.columns and post in df.columns:
                df[f"delta_{v}"] = df[post] - df[pre]
                if not v.startswith("echo"):
                    df[f"ratio_{v}"] = df[post] / df[pre].replace(0, np.nan)
    for v in ["sodium", "potassium", "hco3", "creatinine", "egfr",
              "hemoglobin", "wbc", "ast", "alt", "bili", "lactate", "ph"]:
        pre, post = f"pre_{v}", f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]
    df["delta_cpo"] = df["post_cpo"] - df["pre_cpo"]
    df["recovery_score"] = np.clip((df["delta_cpo"] + 0.5) * 100, 0, 100)
    df["vis_high"] = (df["vis_score"] > 15).astype(int)
    inotrope_cols = ["dopamine", "dobutamine", "epinephrine", "milrinone", "norepinephrine", "vasopressin"]
    df["inotrope_count"] = df[inotrope_cols].gt(0).sum(axis=1)
    scai_map = {"b": 1, "c": 2, "d": 3, "e": 4}
    df["scai_numeric"] = df["scai_stage"].astype(str).str.lower().map(scai_map)
    df["shock_cause_numeric"] = pd.to_numeric(df["cause_of_shock"], errors="coerce")
    df["gender_numeric"] = pd.to_numeric(df["gender"], errors="coerce")
    df["race_numeric"] = pd.to_numeric(df["race"], errors="coerce")

    return df


def get_feature_matrix(df, feature_cols, target_col):
    df_model = df.dropna(subset=[target_col])
    y = df_model[target_col].astype(int).values
    X_raw = df_model[feature_cols].copy()
    all_nan = X_raw.columns[X_raw.isna().all()].tolist()
    X_raw = X_raw.drop(columns=[c for c in all_nan if c in X_raw.columns])
    valid = X_raw.columns.tolist()
    imputer = SimpleImputer(strategy="median")
    X_imp = imputer.fit_transform(X_raw)
    scaler = StandardScaler()
    X_scl = scaler.fit_transform(X_imp)
    return X_scl, y, valid


def evaluate_subset(X, y):
    """5-fold CV AUC for both LR and RF."""
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    lr = LogisticRegression(max_iter=1000, class_weight="balanced", random_state=RANDOM_STATE)
    rf = RandomForestClassifier(n_estimators=200, max_depth=6, class_weight="balanced", random_state=RANDOM_STATE)
    lr_auc = cross_val_score(lr, X, y, cv=cv, scoring="roc_auc").mean()
    rf_auc = cross_val_score(rf, X, y, cv=cv, scoring="roc_auc").mean()
    return {"LR_AUC": round(lr_auc, 4), "RF_AUC": round(rf_auc, 4)}


def main():
    print("=" * 60)
    print("Feature Subset Comparison")
    print("=" * 60)

    print("\n[1/4] Loading consensus rankings...")
    consensus = pd.read_csv(CONSENSUS_PATH, index_col=0)
    top_features = consensus.index.tolist()
    print(f"  Loaded {len(top_features)} ranked features")

    print("\n[2/4] Loading and preparing data...")
    df = load_and_prepare()

    # Get all numeric features (same logic as ml_pipeline.py)
    exclude = {"mrn", "first_name", "last_name", "general_notes", "date_of_implant",
               "cohort_outcome", "cohort_age", "cohort_gender", "indication",
               "physician", "rhc_prior_72h", "rhc_timing_days", "support_days",
               "target_survival", "target_escalation", "target_rv_dysfunction",
               "pre_septal_flattening", "pre_atrial_bowing",
               "post_septal_flattening", "post_atrial_bowing"}
    all_numeric = [c for c in df.columns if c not in exclude and pd.api.types.is_numeric_dtype(df[c])]

    targets = [
        ("survival", "target_survival"),
        ("escalation", "target_escalation"),
        ("rv_dysfunction", "target_rv_dysfunction"),
    ]

    results = {}
    for tname, tcol in targets:
        print(f"\n[3/4] Target: {tname}")
        # Full feature set baseline
        X_full, y_full, full_valid = get_feature_matrix(df, all_numeric, tcol)
        full_result = evaluate_subset(X_full, y_full)
        print(f"  Full ({len(full_valid)} features): LR={full_result['LR_AUC']:.4f}  RF={full_result['RF_AUC']:.4f}")

        target_results = {
            "full": {"n_features": len(full_valid), **full_result},
            "subsets": {}
        }

        # Get overlapping top features from consensus
        consensus_available = [f for f in top_features if f in full_valid]
        print(f"  Consensus features available in matrix: {len(consensus_available)}")

        for k in SUBSETS:
            subset_feats = consensus_available[:min(k, len(consensus_available))]
            if len(subset_feats) < 3:
                print(f"  Top-{k}: skipping (only {len(subset_feats)} available)")
                continue

            X_sub = X_full[:, [full_valid.index(f) for f in subset_feats if f in full_valid]]
            sub_result = evaluate_subset(X_sub, y_full)
            target_results["subsets"][str(k)] = {
                "n_features": X_sub.shape[1],
                **sub_result,
                "auc_delta_vs_full_LR": round(sub_result["LR_AUC"] - full_result["LR_AUC"], 4),
                "auc_delta_vs_full_RF": round(sub_result["RF_AUC"] - full_result["RF_AUC"], 4),
            }
            print(f"  Top-{k} ({X_sub.shape[1]} feats): LR={sub_result['LR_AUC']:.4f}  RF={sub_result['RF_AUC']:.4f}  "
                  f"ΔLR={target_results['subsets'][str(k)]['auc_delta_vs_full_LR']:+.4f}  ΔRF={target_results['subsets'][str(k)]['auc_delta_vs_full_RF']:+.4f}")

        results[tname] = target_results

    # Save JSON
    out = OUTPUT_DIR / "feature_subset_comparison.json"
    with open(out, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\n[4/4] Results saved to {out}")

    # Generate comparison plot
    plot_comparison(results)
    # Generate table for dashboard consumption
    generate_summary_table(results)

    print("\nDone!")


def plot_comparison(results: dict):
    n_targets = len(results)
    fig, axes = plt.subplots(1, n_targets, figsize=(6 * n_targets, 5))
    if n_targets == 1:
        axes = [axes]

    colors = {"LR_AUC": "#2196F3", "RF_AUC": "#FF5722"}

    for ax, (tname, tres) in zip(axes, results.items()):
        full_lr = tres["full"]["LR_AUC"]
        full_rf = tres["full"]["RF_AUC"]

        subsets_int = sorted([int(k) for k in tres["subsets"].keys()])
        subset_lr = [tres["subsets"][str(k)]["LR_AUC"] for k in subsets_int]
        subset_rf = [tres["subsets"][str(k)]["RF_AUC"] for k in subsets_int]

        ax.axhline(y=full_lr, color=colors["LR_AUC"], linestyle="--", alpha=0.5, label=f"Full LR ({full_lr:.3f})")
        ax.axhline(y=full_rf, color=colors["RF_AUC"], linestyle="--", alpha=0.5, label=f"Full RF ({full_rf:.3f})")

        ax.plot(subsets_int, subset_lr, "o-", color=colors["LR_AUC"], linewidth=2, markersize=6, label="LR (subset)")
        ax.plot(subsets_int, subset_rf, "s-", color=colors["RF_AUC"], linewidth=2, markersize=6, label="RF (subset)")

        tname_label = tname.replace("_", " ").title()
        ax.set_xlabel("Top-K Features (by consensus)", fontsize=11)
        ax.set_ylabel("5-Fold CV AUC", fontsize=11)
        ax.set_title(f"Target: {tname_label}", fontsize=13, fontweight="bold")
        ax.legend(fontsize=9)
        ax.set_xticks(subsets_int)
        ax.grid(True, alpha=0.3)
        ax.set_ylim(0.3, 1.0)

    plt.suptitle("Feature Subset Performance Comparison", fontsize=15, fontweight="bold", y=1.02)
    plt.tight_layout()
    plot_path = OUTPUT_DIR / "feature_subset_comparison.png"
    plt.savefig(plot_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Comparison plot saved to {plot_path}")


def generate_summary_table(results: dict):
    """Generate a CSV summary table."""
    rows = []
    for tname, tres in results.items():
        rows.append({
            "target": tname, "subset": "full",
            "n_features": tres["full"]["n_features"],
            "LR_AUC": tres["full"]["LR_AUC"],
            "RF_AUC": tres["full"]["RF_AUC"],
            "delta_LR": 0, "delta_RF": 0,
        })
        for k, sub in sorted(tres["subsets"].items(), key=lambda x: int(x[0])):
            rows.append({
                "target": tname, "subset": f"top_{k}",
                "n_features": sub["n_features"],
                "LR_AUC": sub["LR_AUC"],
                "RF_AUC": sub["RF_AUC"],
                "delta_LR": sub["auc_delta_vs_full_LR"],
                "delta_RF": sub["auc_delta_vs_full_RF"],
            })

    summary = pd.DataFrame(rows)
    summary_path = OUTPUT_DIR / "feature_subset_summary.csv"
    summary.to_csv(summary_path, index=False)
    print(f"  Summary table saved to {summary_path}")


if __name__ == "__main__":
    main()
