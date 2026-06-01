"""
Impella Analytics ML Pipeline
=============================
Trains and evaluates ML models for three clinical prediction targets:
1. Survival (binary: survived vs expired)
2. MCS Escalation (binary: ECMO/LVAD/Transplant/Arrest vs none)
3. RV Dysfunction (binary: composite clinical criteria)

Data sources:
- Patient Data sheet (67 patients, 135 variables)
- Cohort sheet (112 patients, merged by MRN)
"""

import json
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    roc_auc_score, confusion_matrix, roc_curve
)
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier

import joblib

warnings.filterwarnings("ignore")

# Import model weights exporter for automatic JS inference update
from export_model_weights import export as export_model_weights

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_PATH = Path("Impella_MK.xlsx")
OUTPUT_DIR = Path("ml_output")
OUTPUT_DIR.mkdir(exist_ok=True)

RANDOM_STATE = 42
N_SPLITS = 5

# Row mapping for Patient Data sheet (1-based Excel row numbers)
PATIENT_DATA_ROWS = {
    "general_notes": 2,
    "first_name": 3,
    "last_name": 4,
    "mrn": 5,
    "date_of_implant": 6,
    "age": 7,
    "weight_kg": 8,
    "height_cm": 9,
    "gender": 10,
    "race": 11,
    "cause_of_shock": 12,
    "scai_stage": 13,
    "days_between_rhc_and_impella": 14,
    # Pre-implant RHC (rows 17-35)
    "pre_ra": 17,
    "pre_rvsp": 18,
    "pre_rvdp": 19,
    "pre_pasp": 20,
    "pre_padp": 21,
    "pre_map": 27,
    "pre_pcwp": 23,
    "pre_pvr": 24,
    "pre_sbp": 25,
    "pre_dbp": 26,
    "pre_hr": 28,
    "pre_tdco": 29,
    "pre_sv": 30,
    "pre_pa_o2": 31,
    "pre_sp_o2": 32,
    "pre_papi": 33,
    "pre_cpo": 34,
    "pre_rv_cpo": 35,
    # Post-implant RHC (rows 40-58)
    "post_ra": 40,
    "post_rvsp": 41,
    "post_rvdp": 42,
    "post_pasp": 43,
    "post_padp": 44,
    "post_map": 50,
    "post_pcwp": 46,
    "post_pvr": 47,
    "post_sbp": 48,
    "post_dbp": 49,
    "post_hr": 51,
    "post_tdco": 52,
    "post_sv": 53,
    "post_pa_o2": 54,
    "post_sp_o2": 55,
    "post_papi": 56,
    "post_cpo": 57,
    "post_rv_cpo": 58,
    # Echo pre (rows 67-75)
    "pre_rvedd": 67,
    "pre_tapse": 68,
    "pre_rv_s": 69,
    "pre_rv_fs": 70,
    "pre_tr_severity": 71,
    "pre_echo_pasp": 72,
    "pre_lvedd": 73,
    "pre_septal_flattening": 74,
    "pre_atrial_bowing": 75,
    # Echo post (rows 78-86)
    "post_rvedd": 78,
    "post_tapse": 79,
    "post_rv_s": 80,
    "post_rv_fs": 81,
    "post_tr_severity": 82,
    "post_echo_pasp": 83,
    "post_lvedd": 84,
    "post_septal_flattening": 85,
    "post_atrial_bowing": 86,
    # Labs pre (rows 88-98)
    "pre_sodium": 88,
    "pre_potassium": 89,
    "pre_hco3": 90,
    "pre_creatinine": 91,
    "pre_egfr": 92,
    "pre_hemoglobin": 93,
    "pre_wbc": 94,
    "pre_ast": 95,
    "pre_alt": 96,
    "pre_bili": 97,
    "pre_lactate": 98,
    "pre_ph": 99,
    # Labs post (rows 102-113)
    "post_sodium": 102,
    "post_potassium": 103,
    "post_hco3": 104,
    "post_creatinine": 105,
    "post_egfr": 106,
    "post_hemoglobin": 107,
    "post_wbc": 108,
    "post_ast": 109,
    "post_alt": 110,
    "post_bili": 111,
    "post_lactate": 112,
    "post_ph": 113,
    # Inotropes (rows 116-122)
    "dopamine": 116,
    "dobutamine": 117,
    "epinephrine": 118,
    "milrinone": 119,
    "norepinephrine": 120,
    "vasopressin": 121,
    "vis_score": 122,
    # Diuretics pre (rows 125-126)
    "pre_furosemide": 125,
    "pre_augmentation": 126,
    # Diuretics post (rows 129-130)
    "post_furosemide": 129,
    "post_augmentation": 130,
    # Impella (rows 133-134)
    "impella_performance": 133,
    "impella_flow": 134,
    # Outcomes (rows 138-141)
    "renal_failure": 138,
    "intubation": 139,
    "mcs_escalation": 140,
    "outcome": 141,
    # PV Loop (rows 144-154)
    "ees": 144,
    "ea": 145,
    "ees_ea": 146,
    "esp": 147,
    "edp": 148,
    "pmax": 149,
    "esv": 150,
    "edv": 151,
    "pv_sv": 152,
    "dp_dt_max": 153,
    "dp_dt_min": 154,
}

# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_patient_data(path: Path) -> pd.DataFrame:
    """Load Patient Data sheet and return one row per patient."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Patient Data"]

    rows = []
    for col in range(2, ws.max_column + 1):
        mrn = ws.cell(row=5, column=col).value
        if not mrn:
            continue
        record = {"mrn": str(mrn).strip()}
        for key, row_idx in PATIENT_DATA_ROWS.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, str):
                val = val.strip()
                if val.lower() in ("n/a", "na", "", "none"):
                    val = np.nan
            record[key] = val
        rows.append(record)

    df = pd.DataFrame(rows)
    numeric_keys = [k for k in PATIENT_DATA_ROWS.keys() if k not in (
        "general_notes", "first_name", "last_name", "mrn", "date_of_implant"
    )]
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")
    return df


def load_cohort(path: Path) -> pd.DataFrame:
    """Load Cohort sheet."""
    df = pd.read_excel(path, sheet_name="Cohort")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    rename_map = {
        "MRN": "mrn",
        "Outcome": "cohort_outcome",
        "Age ": "cohort_age",
        "Gender": "cohort_gender",
        "Indication for Use": "indication",
        "Duration of Support Time (Days)": "support_days",
        "Physician Name": "physician",
        "RHC prior to implant (<72 hours)": "rhc_prior_72h",
        "RHC timing (days prior)": "rhc_timing_days",
    }
    df = df.rename(columns=rename_map)
    df["mrn"] = df["mrn"].astype(str).str.strip()
    return df


# ---------------------------------------------------------------------------
# Target Engineering
# ---------------------------------------------------------------------------

def build_targets(df: pd.DataFrame) -> pd.DataFrame:
    """Create three prediction targets."""
    # 1. Survival: 1 = expired, 0 = survived
    df["target_survival"] = (df["cohort_outcome"].str.lower() == "expired").astype(int)

    # 2. MCS Escalation: 1 = ECMO/LVAD/Transplant/Arrest after impella
    notes = df["general_notes"].fillna("").str.lower()
    escalation_keywords = ["ecmo", "lvad", "transplant", "arrest", "rv failure", "rvf"]
    has_escalation_keyword = notes.apply(
        lambda x: any(kw in x for kw in escalation_keywords)
    )
    mcs_flag = df["mcs_escalation"].fillna(0) > 0
    df["target_escalation"] = ((has_escalation_keyword) | mcs_flag).astype(int)

    # 3. RV Dysfunction: composite clinical criteria
    rv_dys = pd.Series(False, index=df.index)
    rv_dys |= df["post_papi"] < 1.0
    rv_dys |= df["post_ra"] > 20
    rv_dys |= df["post_tapse"] < 1.6
    rv_dys |= df["post_rv_s"] < 9.5
    rv_dys |= df["post_rv_cpo"] < (df["pre_rv_cpo"] * 0.7)
    rv_dys |= notes.str.contains("rv failure|rv dysfunction|rvf", regex=True, na=False)
    df["target_rv_dysfunction"] = rv_dys.astype(int)

    return df


# ---------------------------------------------------------------------------
# Feature Engineering
# ---------------------------------------------------------------------------

def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    """Create delta, ratio, and composite features."""
    df = df.copy()

    df["bmi"] = df["weight_kg"] / ((df["height_cm"] / 100) ** 2)

    # Delta features (post - pre)
    rhc_vars = [
        "ra", "rvsp", "rvdp", "pasp", "padp", "map", "pcwp", "pvr",
        "sbp", "dbp", "hr", "tdco", "sv", "pa_o2", "sp_o2", "papi",
        "cpo", "rv_cpo",
    ]
    for v in rhc_vars:
        pre = f"pre_{v}"
        post = f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]
            df[f"ratio_{v}"] = df[post] / df[pre].replace(0, np.nan)

    echo_vars = ["rvedd", "tapse", "rv_s", "rv_fs", "tr_severity", "echo_pasp", "lvedd"]
    for v in echo_vars:
        pre = f"pre_{v}"
        post = f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]
            df[f"ratio_{v}"] = df[post] / df[pre].replace(0, np.nan)

    lab_vars = ["sodium", "potassium", "hco3", "creatinine", "egfr",
                "hemoglobin", "wbc", "ast", "alt", "bili", "lactate", "ph"]
    for v in lab_vars:
        pre = f"pre_{v}"
        post = f"post_{v}"
        if pre in df.columns and post in df.columns:
            df[f"delta_{v}"] = df[post] - df[pre]

    df["delta_cpo"] = df["post_cpo"] - df["pre_cpo"]
    df["recovery_score"] = np.clip((df["delta_cpo"] + 0.5) * 100, 0, 100)

    df["vis_high"] = (df["vis_score"] > 15).astype(int)

    inotrope_cols = ["dopamine", "dobutamine", "epinephrine", "milrinone", "norepinephrine", "vasopressin"]
    df["inotrope_count"] = df[inotrope_cols].gt(0).sum(axis=1)

    scai_map = {"b": 1, "c": 2, "d": 3, "e": 4}
    df["scai_numeric"] = df["scai_stage"].astype(str).str.lower().map(scai_map)

    df["shock_cause_numeric"] = pd.to_numeric(df["cause_of_shock"], errors="coerce")
    df["gender_numeric"] = pd.to_numeric(df["gender"], errors="coerce")
    df["race_numeric"] = pd.to_numeric(df["race"], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# Preprocessing
# ---------------------------------------------------------------------------

def prepare_features(df: pd.DataFrame, feature_cols: list) -> tuple:
    """Return X, imputer, scaler, valid_feature_cols."""
    X = df[feature_cols].copy()
    # Drop columns that are entirely NaN (SimpleImputer would drop them silently)
    all_nan_cols = X.columns[X.isna().all()].tolist()
    if all_nan_cols:
        print(f"    Dropping {len(all_nan_cols)} all-NaN columns: {all_nan_cols[:5]}{'...' if len(all_nan_cols) > 5 else ''}")
        X = X.drop(columns=all_nan_cols)
    valid_cols = X.columns.tolist()
    imputer = SimpleImputer(strategy="median")
    X_imputed = imputer.fit_transform(X)
    X = pd.DataFrame(X_imputed, columns=valid_cols, index=X.index)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    X = pd.DataFrame(X_scaled, columns=valid_cols, index=X.index)
    return X, imputer, scaler, valid_cols


# ---------------------------------------------------------------------------
# Model Training & Evaluation
# ---------------------------------------------------------------------------

def evaluate_model(model, X: pd.DataFrame, y: pd.Series, cv: StratifiedKFold) -> dict:
    y_proba = cross_val_predict(model, X, y, cv=cv, method="predict_proba")[:, 1]
    y_pred = (y_proba >= 0.5).astype(int)
    return {
        "auc": roc_auc_score(y, y_proba),
        "accuracy": accuracy_score(y, y_pred),
        "precision": precision_score(y, y_pred, zero_division=0),
        "recall": recall_score(y, y_pred, zero_division=0),
        "f1": f1_score(y, y_pred, zero_division=0),
        "y_true": y.values,
        "y_proba": y_proba,
        "y_pred": y_pred,
    }


def train_all_models(X: pd.DataFrame, y: pd.Series, target_name: str) -> dict:
    cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    models = {
        "LogisticRegression": LogisticRegression(max_iter=1000, random_state=RANDOM_STATE, class_weight="balanced"),
        "RandomForest": RandomForestClassifier(n_estimators=200, max_depth=6, random_state=RANDOM_STATE, class_weight="balanced"),
    }
    results = {}
    for name, model in models.items():
        print(f"  Training {name} ...")
        metrics = evaluate_model(model, X, y, cv)
        results[name] = metrics
        print(f"    AUC={metrics['auc']:.3f}  F1={metrics['f1']:.3f}  Acc={metrics['accuracy']:.3f}")
    return results


def plot_roc_curves(results: dict, target_name: str, output_dir: Path):
    plt.figure(figsize=(8, 6))
    for name, metrics in results.items():
        fpr, tpr, _ = roc_curve(metrics["y_true"], metrics["y_proba"])
        plt.plot(fpr, tpr, label=f"{name} (AUC={metrics['auc']:.3f})")
    plt.plot([0, 1], [0, 1], "k--", label="Chance")
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title(f"ROC Curves — {target_name}")
    plt.legend(loc="lower right")
    plt.tight_layout()
    plt.savefig(output_dir / f"roc_{target_name}.png", dpi=150)
    plt.close()


def plot_confusion_matrices(results: dict, target_name: str, output_dir: Path):
    n = len(results)
    fig, axes = plt.subplots(1, n, figsize=(5 * n, 4))
    if n == 1:
        axes = [axes]
    for ax, (name, metrics) in zip(axes, results.items()):
        cm = confusion_matrix(metrics["y_true"], metrics["y_pred"])
        sns.heatmap(cm, annot=True, fmt="d", cmap="Blues", ax=ax, cbar=False)
        ax.set_title(name)
        ax.set_xlabel("Predicted")
        ax.set_ylabel("Actual")
    plt.suptitle(f"Confusion Matrices — {target_name}")
    plt.tight_layout()
    plt.savefig(output_dir / f"cm_{target_name}.png", dpi=150)
    plt.close()


# ---------------------------------------------------------------------------
# SHAP Explainability
# ---------------------------------------------------------------------------

def shap_summary(model, X: pd.DataFrame, feature_names: list, target_name: str, output_dir: Path):
    try:
        import shap
        if hasattr(model, "estimators_"):
            explainer = shap.TreeExplainer(model)
        else:
            explainer = shap.KernelExplainer(model.predict_proba, X)
        shap_values = explainer.shap_values(X)
        sv = shap_values[1] if isinstance(shap_values, list) else shap_values
        plt.figure(figsize=(10, 6))
        shap.summary_plot(sv, X, feature_names=feature_names, show=False)
        plt.title(f"SHAP Summary — {target_name}")
        plt.tight_layout()
        plt.savefig(output_dir / f"shap_{target_name}.png", dpi=150)
        plt.close()
    except Exception as e:
        print(f"  SHAP plotting skipped: {e}")


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_lr_json(model: LogisticRegression, imputer, scaler, feature_names: list, target_name: str, output_dir: Path):
    """Export logistic regression coefficients as JSON for Node.js consumption."""
    payload = {
        "model_type": "LogisticRegression",
        "target": target_name,
        "feature_names": feature_names,
        "coef": model.coef_[0].tolist(),
        "intercept": float(model.intercept_[0]),
        "imputer_statistics": imputer.statistics_.tolist(),
        "scaler_mean": scaler.mean_.tolist(),
        "scaler_scale": scaler.scale_.tolist(),
    }
    path = output_dir / f"model_{target_name}_lr.json"
    with open(path, "w") as f:
        json.dump(payload, f, indent=2)
    print(f"  Exported LR coefficients to {path}")


def export_sklearn_joblib(model, imputer, scaler, feature_names: list, target_name: str, output_dir: Path):
    """Export full sklearn model artifact via joblib for Python consumption."""
    artifact = {
        "model": model,
        "imputer": imputer,
        "scaler": scaler,
        "feature_names": feature_names,
        "target_name": target_name,
    }
    path = output_dir / f"model_{target_name}.joblib"
    joblib.dump(artifact, path)
    print(f"  Exported joblib artifact to {path}")


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------

def generate_report(all_results: dict, feature_names: list, output_dir: Path):
    lines = [
        "# Impella Analytics ML Model Report",
        "",
        "## Overview",
        "",
        "This report summarizes machine-learning models trained to predict three clinical outcomes for Impella-supported patients.",
        "",
        "## Targets",
        "",
        "1. **Survival** — Whether the patient expired during the hospitalization.",
        "2. **MCS Escalation** — Whether the patient required ECMO, LVAD, transplant, or had a post-implant arrest.",
        "3. **RV Dysfunction** — Composite clinical criteria (PAPI < 1.0, RA > 20, TAPSE < 1.6, RV S' < 9.5, RV-CPO drop > 30%, or explicit RV failure in notes).",
        "",
        "## Feature Set",
        "",
        f"Total engineered features: **{len(feature_names)}**",
        "",
        "### Categories",
        "- Demographics & baseline (age, BMI, gender, race, SCAI stage, cause of shock)",
        "- Pre-implant RHC hemodynamics",
        "- Post-implant RHC hemodynamics",
        "- Delta & ratio features (post − pre, post / pre)",
        "- Echo metrics (pre & post)",
        "- Laboratory values (pre & post)",
        "- Inotrope & diuretic data",
        "- Impella settings (performance level, flow)",
        "- PV loop mechanics (Ees, Ea, Ees/Ea)",
        "",
        "## Cross-Validation Results",
        "",
    ]

    for target, results in all_results.items():
        lines.append(f"### {target}")
        lines.append("")
        lines.append("| Model | AUC | Accuracy | Precision | Recall | F1 |")
        lines.append("|-------|-----|----------|-----------|--------|----|")
        for name, metrics in results.items():
            lines.append(
                f"| {name} | {metrics['auc']:.3f} | {metrics['accuracy']:.3f} | "
                f"{metrics['precision']:.3f} | {metrics['recall']:.3f} | {metrics['f1']:.3f} |"
            )
        best = max(results.items(), key=lambda kv: kv[1]["auc"])
        lines.append("")
        lines.append(f"**Best model:** {best[0]} (AUC = {best[1]['auc']:.3f})")
        lines.append("")

    lines.extend([
        "## Clinical Interpretation",
        "",
        "- **AUC > 0.75** is considered clinically useful for risk stratification.",
        "- **AUC 0.65–0.75** provides moderate discrimination and may be useful as a screening tool.",
        "- Models with low recall on the positive class may miss high-risk patients; consider lowering the decision threshold in production.",
        "",
        "## Limitations",
        "",
        "- Small sample size (67–112 patients) limits model generalizability.",
        "- Missing data is imputed with median values; this may obscure true clinical trajectories.",
        "- RV dysfunction target is partially derived from echo data, which is missing for some patients.",
        "- The models have not been externally validated on an independent cohort.",
        "",
        "## Next Steps",
        "",
        "1. External validation on a separate institutional cohort.",
        "2. Collect more complete echo and lab data to reduce imputation.",
        "3. Implement dynamic threshold tuning based on clinical cost of false negatives.",
        "4. Integrate the best model into the Node.js dashboard via the exported JSON coefficients.",
        "",
    ])

    report_path = output_dir / "model_report.md"
    with open(report_path, "w") as f:
        f.write("\n".join(lines))
    print(f"\nReport written to {report_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("Impella Analytics ML Pipeline")
    print("=" * 60)

    print("\n[1/6] Loading data...")
    df_pd = load_patient_data(DATA_PATH)
    df_cohort = load_cohort(DATA_PATH)
    print(f"  Patient Data: {len(df_pd)} patients")
    print(f"  Cohort: {len(df_cohort)} patients")

    df = df_pd.merge(df_cohort, on="mrn", how="outer")
    print(f"  Merged dataset: {len(df)} patients")

    print("\n[2/6] Engineering targets...")
    df = build_targets(df)
    print(f"  Survival (expired): {df['target_survival'].sum()} / {len(df)}")
    print(f"  Escalation: {df['target_escalation'].sum()} / {len(df)}")
    print(f"  RV Dysfunction: {df['target_rv_dysfunction'].sum()} / {len(df)}")

    print("\n[3/6] Engineering features...")
    df = engineer_features(df)

    exclude = {"mrn", "first_name", "last_name", "general_notes", "date_of_implant",
               "cohort_outcome", "cohort_age", "cohort_gender", "indication",
               "physician", "rhc_prior_72h", "rhc_timing_days", "support_days",
               "target_survival", "target_escalation", "target_rv_dysfunction",
               "pre_septal_flattening", "pre_atrial_bowing",
               "post_septal_flattening", "post_atrial_bowing"}
    feature_cols = [c for c in df.columns if c not in exclude and pd.api.types.is_numeric_dtype(df[c])]
    print(f"  Using {len(feature_cols)} numeric features")

    df_model = df.dropna(subset=["target_survival", "target_escalation", "target_rv_dysfunction"], how="all")
    print(f"  Model-ready rows: {len(df_model)}")

    targets = {
        "survival": "target_survival",
        "escalation": "target_escalation",
        "rv_dysfunction": "target_rv_dysfunction",
    }

    all_results = {}
    best_models = {}

    print("\n[4/6] Training models...")
    for target_name, target_col in targets.items():
        df_target = df_model.dropna(subset=[target_col])
        y = df_target[target_col].astype(int)
        if y.nunique() < 2:
            print(f"  Skipping {target_name}: only one class present.")
            continue
        if len(y) < 10:
            print(f"  Skipping {target_name}: insufficient samples ({len(y)}).")
            continue

        print(f"\n  Target: {target_name} (n={len(y)}, pos={y.sum()})")
        X, imputer, scaler, valid_cols = prepare_features(df_target, feature_cols)

        results = train_all_models(X, y, target_name)
        all_results[target_name] = results

        best_name = max(results, key=lambda k: results[k]["auc"])
        best_models[target_name] = best_name
        print(f"  Best model: {best_name}")

        best_clf = {
            "LogisticRegression": LogisticRegression(max_iter=1000, random_state=RANDOM_STATE, class_weight="balanced"),
            "RandomForest": RandomForestClassifier(n_estimators=200, max_depth=6, random_state=RANDOM_STATE, class_weight="balanced"),
        }[best_name]
        best_clf.fit(X, y)

        plot_roc_curves(results, target_name, OUTPUT_DIR)
        plot_confusion_matrices(results, target_name, OUTPUT_DIR)
        shap_summary(best_clf, X, valid_cols, target_name, OUTPUT_DIR)

        if isinstance(best_clf, LogisticRegression):
            export_lr_json(best_clf, imputer, scaler, valid_cols, target_name, OUTPUT_DIR)
        export_sklearn_joblib(best_clf, imputer, scaler, valid_cols, target_name, OUTPUT_DIR)

    print("\n[5/6] Generating report...")
    generate_report(all_results, feature_cols, OUTPUT_DIR)

    print("\n[6/6] Saving feature metadata...")
    meta = {
        "feature_names": feature_cols,
        "targets": list(targets.keys()),
        "best_models": best_models,
        "n_patients": len(df_model),
    }
    with open(OUTPUT_DIR / "model_metadata.json", "w") as f:
        json.dump(meta, f, indent=2)

    print("\n[7/6] Exporting model weights for JS inference...")
    export_model_weights()

    print("\nPipeline complete.")


if __name__ == "__main__":
    main()
