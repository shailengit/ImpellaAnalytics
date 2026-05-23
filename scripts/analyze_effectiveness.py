"""
Impella Effectiveness Analysis
===============================
Reads Impella_MK.xlsx and produces a multi-dimensional analysis of Impella
hemodynamic support effectiveness. Exports structured JSON for the dashboard
and generates a self-contained HTML report.

Analysis dimensions:
1. Survival & Outcomes — mortality by indication, demographics
2. Hemodynamic Response — pre/post delta with statistical tests
3. Organ Function & Labs — lactate clearance, renal, hemolysis
4. Responder Profiling — multi-dimensional responder characterization
5. Ventricular Mechanics — Ees/Ea coupling, ESP, EDP
6. ML Model Cross-Reference — consensus features mapped to clinical findings
"""

import json
import warnings
import base64
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import mannwhitneyu, chi2_contingency
from io import BytesIO

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_PATH = Path("Impella_MK.xlsx")
OUTPUT_JSON = Path("public/effectiveness-data.json")
OUTPUT_REPORT = Path("public/effectiveness-report.html")
OUTPUT_DIR = Path("public")
OUTPUT_DIR.mkdir(exist_ok=True)

# Row mapping for Patient Data sheet
PATIENT_DATA_ROWS = {
    "general_notes": 2, "first_name": 3, "last_name": 4, "mrn": 5,
    "date_of_implant": 6, "age": 7, "weight_kg": 8, "height_cm": 9,
    "gender": 10, "race": 11, "cause_of_shock": 12, "scai_stage": 13,
    "days_between_rhc_and_impella": 14,
    "pre_ra": 17, "pre_rvsp": 18, "pre_rvdp": 19, "pre_pasp": 20,
    "pre_padp": 21, "pre_map": 27, "pre_pcwp": 23, "pre_pvr": 24,
    "pre_sbp": 25, "pre_dbp": 26, "pre_hr": 28, "pre_tdco": 29,
    "pre_sv": 30, "pre_pa_o2": 31, "pre_sp_o2": 32, "pre_papi": 33,
    "pre_cpo": 34, "pre_rv_cpo": 35,
    "post_ra": 40, "post_rvsp": 41, "post_rvdp": 42, "post_pasp": 43,
    "post_padp": 44, "post_map": 50, "post_pcwp": 46, "post_pvr": 47,
    "post_sbp": 48, "post_dbp": 49, "post_hr": 51, "post_tdco": 52,
    "post_sv": 53, "post_pa_o2": 54, "post_sp_o2": 55, "post_papi": 56,
    "post_cpo": 57, "post_rv_cpo": 58,
    "pre_rvedd": 67, "pre_tapse": 68, "pre_rv_s": 69, "pre_rv_fs": 70,
    "pre_tr_severity": 71, "pre_echo_pasp": 72, "pre_lvedd": 73,
    "post_rvedd": 78, "post_tapse": 79, "post_rv_s": 80, "post_rv_fs": 81,
    "post_tr_severity": 82, "post_echo_pasp": 83, "post_lvedd": 84,
    "pre_sodium": 88, "pre_potassium": 89, "pre_hco3": 90,
    "pre_creatinine": 91, "pre_egfr": 92, "pre_hemoglobin": 93, "pre_wbc": 94,
    "pre_ast": 95, "pre_alt": 96, "pre_bili": 97, "pre_lactate": 98, "pre_ph": 99,
    "post_sodium": 102, "post_potassium": 103, "post_hco3": 104,
    "post_creatinine": 105, "post_egfr": 106, "post_hemoglobin": 107,
    "post_wbc": 108, "post_ast": 109, "post_alt": 110, "post_bili": 111,
    "post_lactate": 112, "post_ph": 113,
    "dopamine": 116, "dobutamine": 117, "epinephrine": 118, "milrinone": 119,
    "norepinephrine": 120, "vasopressin": 121, "vis_score": 122,
    "pre_furosemide": 125, "post_furosemide": 129,
    "impella_performance": 133, "impella_flow": 134,
    "renal_failure": 138, "intubation": 139, "mcs_escalation": 140, "outcome": 141,
    "ees": 144, "ea": 145, "ees_ea": 146, "esp": 147, "edp": 148,
    "pmax": 149, "esv": 150, "edv": 151, "pv_sv": 152, "dp_dt_max": 153, "dp_dt_min": 154,
}

# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_patient_data(path: Path) -> pd.DataFrame:
    """Load Patient Data sheet, return one row per patient."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Patient Data"]
    rows = []
    for col in range(2, ws.max_column + 1):
        mrn = ws.cell(row=5, column=col).value
        if not mrn:
            continue
        record = {"mrn": str(mrn).strip()}
        for key, row_idx in PATIENT_DATA_ROWS.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None and isinstance(val, str):
                val = val.strip()
                if val.lower() in ("n/a", "na", "", "none"):
                    val = np.nan
            record[key] = val
        rows.append(record)
    df = pd.DataFrame(rows)
    numeric_keys = [k for k in PATIENT_DATA_ROWS if k not in (
        "general_notes", "first_name", "last_name", "mrn", "date_of_implant"
    )]
    for k in numeric_keys:
        df[k] = pd.to_numeric(df[k], errors="coerce")
    return df


def load_cohort(path: Path) -> pd.DataFrame:
    """Load Cohort sheet."""
    df = pd.read_excel(path, sheet_name="Cohort")
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    rename_map = {
        "MRN": "mrn", "Outcome": "cohort_outcome", "Indication for Use": "indication",
        "Duration of Support Time (Days)": "support_days",
        "Physician Name": "physician",
    }
    df = df.rename(columns=rename_map)
    df["mrn"] = df["mrn"].astype(str).str.strip()
    df["support_days"] = pd.to_numeric(df["support_days"], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Analysis Functions
# ---------------------------------------------------------------------------

def compute_survival_analysis(df: pd.DataFrame, cohort: pd.DataFrame) -> dict:
    """Survival & outcomes by indication, demographics."""
    merged = df.merge(cohort, on="mrn", how="inner")
    results = {}

    # By indication
    indication_groups = merged.groupby("indication")
    indication_stats = []
    for ind, grp in indication_groups:
        n = len(grp)
        expired = grp["cohort_outcome"].str.lower().eq("expired").sum()
        rate = round(expired / n * 100, 1) if n > 0 else 0
        indication_stats.append({
            "indication": ind,
            "total": n,
            "survived": int(n - expired),
            "expired": int(expired),
            "mortalityRate": rate,
        })
    results["byIndication"] = indication_stats

    # Overall
    total = len(merged)
    expired = merged["cohort_outcome"].str.lower().eq("expired").sum()
    results["overall"] = {
        "total": int(total),
        "survived": int(total - expired),
        "expired": int(expired),
        "mortalityRate": round(expired / total * 100, 1) if total > 0 else 0,
    }

    # Support duration by outcome
    sup = merged.dropna(subset=["cohort_outcome", "support_days"])
    survived_sup = sup[sup["cohort_outcome"].str.lower() == "survived"]["support_days"]
    expired_sup = sup[sup["cohort_outcome"].str.lower() == "expired"]["support_days"]
    results["supportDuration"] = {
        "survived": {
            "mean": round(survived_sup.mean(), 1) if len(survived_sup) > 0 else 0,
            "median": round(survived_sup.median(), 1) if len(survived_sup) > 0 else 0,
        },
        "expired": {
            "mean": round(expired_sup.mean(), 1) if len(expired_sup) > 0 else 0,
            "median": round(expired_sup.median(), 1) if len(expired_sup) > 0 else 0,
        },
    }

    return results


def compute_hemodynamic_analysis(df: pd.DataFrame, cohort: pd.DataFrame) -> dict:
    """Pre/post hemodynamic delta analysis with statistical tests."""
    merged = df.merge(cohort, on="mrn", how="inner")
    merged["outcome_label"] = merged["cohort_outcome"].str.lower()

    hemodynamic_pairs = [
        ("cpo", "CPO", "W"),
        ("tdco", "TDCO", "L/min"),
        ("rv_cpo", "RV-CPO", "W"),
        ("pcwp", "PCWP", "mmHg"),
        ("map", "MAP", "mmHg"),
        ("ra", "RA", "mmHg"),
        ("papi", "PAPI", ""),
        ("pvr", "PVR", "WU"),
        ("hr", "HR", "bpm"),
        ("sv", "SV", "mL"),
    ]

    deltas = []
    stat_tests = []

    for key, label, unit in hemodynamic_pairs:
        pre_key = f"pre_{key}"
        post_key = f"post_{key}"
        if pre_key not in df.columns or post_key not in df.columns:
            continue

        d = merged.copy()
        d["delta"] = d[post_key] - d[pre_key]
        d["metric"] = label
        d["unit"] = unit

        for _, row in d.iterrows():
            deltas.append({
                "patientId": row.get("first_name", ""),
                "metric": label,
                "unit": unit,
                "pre": round(row[pre_key], 3) if pd.notna(row[pre_key]) else None,
                "post": round(row[post_key], 3) if pd.notna(row[post_key]) else None,
                "delta": round(row["delta"], 3) if pd.notna(row["delta"]) else None,
                "outcome": row["outcome_label"] if row["outcome_label"] in ("survived", "expired") else "unknown",
            })

        # Statistical test: Mann-Whitney comparing delta between survived vs expired
        surv = d[d["outcome_label"] == "survived"]["delta"].dropna()
        exp = d[d["outcome_label"] == "expired"]["delta"].dropna()
        if len(surv) > 2 and len(exp) > 2:
            stat, pval = mannwhitneyu(surv, exp, alternative="two-sided")
            stat_tests.append({
                "metric": label,
                "survivedMean": round(surv.mean(), 3),
                "expiredMean": round(exp.mean(), 3),
                "survivedN": len(surv),
                "expiredN": len(exp),
                "pValue": round(float(pval), 4),
                "significant": pval < 0.05,
            })

    # Build pre-computed scatter pairs CPO vs TDCO (paired by row order, avoids duplicate patientId matching issues)
    cpo_entries = [d for d in deltas if d["metric"] == "CPO"]
    tdco_entries = [d for d in deltas if d["metric"] == "TDCO"]
    scatter_data = []
    for cpo, tdco in zip(cpo_entries, tdco_entries):
        scatter_data.append({
            "patientId": cpo["patientId"],
            "cpoDelta": cpo["delta"] if cpo["delta"] is not None else 0,
            "tdcoDelta": tdco["delta"] if tdco["delta"] is not None else 0,
            "outcome": cpo["outcome"],
        })

    return {"deltas": deltas, "statisticalTests": stat_tests, "scatterData": scatter_data}


def compute_lab_analysis(df: pd.DataFrame, cohort: pd.DataFrame) -> dict:
    """Pre/post lab recovery analysis."""
    merged = df.merge(cohort, on="mrn", how="inner")
    merged["outcome_label"] = merged["cohort_outcome"].str.lower()

    lab_pairs = [
        ("lactate", "Lactate", "mmol/L"),
        ("egfr", "eGFR", "mL/min"),
        ("creatinine", "Creatinine", "mg/dL"),
        ("hco3", "HCO3", "mEq/L"),
        ("hemoglobin", "Hemoglobin", "g/dL"),
        ("ast", "AST", "U/L"),
        ("alt", "ALT", "U/L"),
        ("bili", "Bilirubin", "mg/dL"),
    ]

    deltas = []
    stat_tests = []

    for key, label, unit in lab_pairs:
        pre_key = f"pre_{key}"
        post_key = f"post_{key}"
        if pre_key not in df.columns or post_key not in df.columns:
            continue

        d = merged.copy()
        d["delta"] = d[post_key] - d[pre_key]

        for _, row in d.iterrows():
            deltas.append({
                "patientId": row.get("first_name", ""),
                "metric": label,
                "unit": unit,
                "pre": round(row[pre_key], 2) if pd.notna(row[pre_key]) else None,
                "post": round(row[post_key], 2) if pd.notna(row[post_key]) else None,
                "delta": round(row["delta"], 3) if pd.notna(row["delta"]) else None,
                "outcome": row["outcome_label"] if row["outcome_label"] in ("survived", "expired") else "unknown",
            })

        surv = d[d["outcome_label"] == "survived"]["delta"].dropna()
        exp = d[d["outcome_label"] == "expired"]["delta"].dropna()
        if len(surv) > 2 and len(exp) > 2:
            stat, pval = mannwhitneyu(surv, exp, alternative="two-sided")
            stat_tests.append({
                "metric": label,
                "survivedMean": round(surv.mean(), 3),
                "expiredMean": round(exp.mean(), 3),
                "pValue": round(float(pval), 4),
                "significant": pval < 0.05,
            })

    return {"deltas": deltas, "statisticalTests": stat_tests}


def compute_responder_profiles(df: pd.DataFrame, cohort: pd.DataFrame) -> dict:
    """Characterize responders vs non-responders.
    Responder = CPO↑ AND lactate↓ AND survived.
    """
    merged = df.merge(cohort, on="mrn", how="inner")
    merged["outcome_label"] = merged["cohort_outcome"].str.lower()

    # Compute criteria
    merged["cpo_improved"] = (merged["post_cpo"] - merged["pre_cpo"]) > 0
    merged["lactate_cleared"] = (merged["pre_lactate"] - merged["post_lactate"]) > 0
    merged["is_responder"] = (
        merged["cpo_improved"]
        & merged["lactate_cleared"]
        & (merged["outcome_label"] == "survived")
    )

    responders = merged[merged["is_responder"]]
    non_responders = merged[~merged["is_responder"]]

    baseline_vars = [
        ("pre_cpo", "Baseline CPO", "W"),
        ("pre_lactate", "Baseline Lactate", "mmol/L"),
        ("pre_ra", "Baseline RA", "mmHg"),
        ("pre_pcwp", "Baseline PCWP", "mmHg"),
        ("pre_papi", "Baseline PAPI", ""),
        ("pre_tdco", "Baseline TDCO", "L/min"),
        ("pre_egfr", "Baseline eGFR", "mL/min"),
        ("age", "Age", "years"),
        ("pre_map", "Baseline MAP", "mmHg"),
        ("pre_rv_cpo", "Baseline RV-CPO", "W"),
    ]

    profile_vars = []
    for key, label, unit in baseline_vars:
        r_vals = responders[key].dropna()
        nr_vals = non_responders[key].dropna()
        if len(r_vals) < 2 or len(nr_vals) < 2:
            continue
        stat, pval = mannwhitneyu(r_vals, nr_vals, alternative="two-sided")
        profile_vars.append({
            "metric": label,
            "unit": unit,
            "responderMean": round(r_vals.mean(), 3),
            "nonResponderMean": round(nr_vals.mean(), 3),
            "pValue": round(float(pval), 4),
            "significant": pval < 0.05,
        })

    # Individual patient classifications
    patient_list = []
    for _, row in merged.iterrows():
        if pd.isna(row.get("first_name")):
            continue
        patient_list.append({
            "patientId": row.get("first_name", ""),
            "outcome": row["outcome_label"] if row["outcome_label"] in ("survived", "expired") else "unknown",
            "isResponder": bool(row["is_responder"]),
            "cpoDelta": round(row.get("post_cpo", 0) - row.get("pre_cpo", 0), 3),
            "lactateDelta": round(row.get("post_lactate", 0) - row.get("pre_lactate", 0), 3),
            "cpoImproved": bool(row["cpo_improved"]),
            "lactateCleared": bool(row["lactate_cleared"]),
        })

    total_responders = responders["first_name"].notna().sum()
    total_non = non_responders["first_name"].notna().sum()

    return {
        "summary": {
            "totalResponders": int(total_responders),
            "totalNonResponders": int(total_non),
            "responderRate": round(total_responders / (total_responders + total_non) * 100, 1) if (total_responders + total_non) > 0 else 0,
        },
        "profileVariables": profile_vars,
        "patients": patient_list,
    }


def compute_ventricular_mechanics(df: pd.DataFrame, cohort: pd.DataFrame) -> dict:
    """PV Loop metrics pre/post by outcome."""
    merged = df.merge(cohort, on="mrn", how="inner")
    merged["outcome_label"] = merged["cohort_outcome"].str.lower()

    pv_metrics = [
        ("ees", "Ees", "mmHg/mL"),
        ("ea", "Ea", "mmHg/mL"),
        ("ees_ea", "Ees/Ea", ""),
        ("esp", "ESP", "mmHg"),
        ("edp", "EDP", "mmHg"),
        ("pmax", "Pmax", "mmHg"),
        ("esv", "ESV", "mL"),
        ("edv", "EDV", "mL"),
    ]

    deltas = []
    stat_tests = []

    for key, label, unit in pv_metrics:
        if key not in df.columns:
            continue

        d = merged.copy()
        d["value"] = d[key]

        for _, row in d.iterrows():
            deltas.append({
                "patientId": row.get("first_name", ""),
                "metric": label,
                "unit": unit,
                "value": round(row[key], 3) if pd.notna(row[key]) else None,
                "outcome": row["outcome_label"] if row["outcome_label"] in ("survived", "expired") else "unknown",
            })

        surv = d[d["outcome_label"] == "survived"]["value"].dropna()
        exp = d[d["outcome_label"] == "expired"]["value"].dropna()
        if len(surv) > 2 and len(exp) > 2:
            stat, pval = mannwhitneyu(surv, exp, alternative="two-sided")
            stat_tests.append({
                "metric": label,
                "survivedMean": round(surv.mean(), 3),
                "expiredMean": round(exp.mean(), 3),
                "pValue": round(float(pval), 4),
                "significant": pval < 0.05,
            })

    return {"values": deltas, "statisticalTests": stat_tests}


def compute_ml_cross_reference() -> dict:
    """Load consensus features from mortality_feature_consensus.csv if available."""
    consensus_path = Path("mortality_feature_consensus.csv")
    if not consensus_path.exists():
        return {"features": [], "note": "Run mortality feature analysis first"}

    df = pd.read_csv(consensus_path)
    # First column is unnamed in the CSV, contains feature names
    first_col = df.columns[0]
    has_consensus = "consensus_score" in df.columns or "consensus" in df.columns
    consensus_col = "consensus_score" if "consensus_score" in df.columns else ("consensus" if "consensus" in df.columns else None)

    # Map feature names to clinical descriptions
    feature_descriptions = {
        "age": "Patient age at implant",
        "pre_ra": "Baseline right atrial pressure",
        "pre_pcwp": "Baseline pulmonary capillary wedge pressure",
        "pre_cpo": "Baseline cardiac power output",
        "pre_papi": "Baseline pulmonary artery pulsatility index",
        "pre_rv_cpo": "Baseline RV cardiac power output",
        "pre_map": "Baseline mean arterial pressure",
        "pre_tdco": "Baseline thermodilution cardiac output",
        "pre_lactate": "Baseline lactate (organ perfusion)",
        "pre_egfr": "Baseline estimated GFR (renal function)",
        "pre_hemoglobin": "Baseline hemoglobin",
        "pre_ast": "Baseline AST (liver injury)",
        "pre_alt": "Baseline ALT (liver injury)",
        "pre_hco3": "Baseline bicarbonate",
        "post_cpo": "Post-implant cardiac power output",
        "post_papi": "Post-implant PAPI",
        "post_ra": "Post-implant right atrial pressure",
        "post_map": "Post-implant MAP",
        "post_lactate": "Post-implant lactate",
        "delta_cpo": "Change in CPO (post − pre)",
        "delta_ra": "Change in RA pressure",
        "delta_pcwp": "Change in PCWP",
        "delta_lactate": "Change in lactate",
        "ees_ea": "Ventricular-arterial coupling (Ees/Ea)",
        "esp": "End-systolic pressure",
        "edp": "End-diastolic pressure",
        "vis_score": "Vasopressor inotrope score",
        "support_days": "Duration of Impella support",
        "post_lactate": "Post-implant lactate level",
        "post_hemoglobin": "Post-implant hemoglobin",
        "post_hco3": "Post-implant bicarbonate",
        "delta_pcwp": "Change in PCWP",
        "ees_ea": "Ventricular-arterial coupling (Ees/Ea)",
        "map": "Mean arterial pressure",
        "delta_ra": "Change in RA pressure",
        "delta_lactate": "Change in lactate",
        "delta_tdco": "Change in TDCO",
        "delta_cpo": "Change in cardiac power output",
        "post_papi": "Post-implant PAPI",
        "pre_hco3": "Baseline bicarbonate",
        "pre_creatinine": "Baseline creatinine",
        "pre_cpo": "Baseline cardiac power output",
        "post_ra": "Post-implant right atrial pressure",
        "pre_lactate": "Baseline lactate",
        "delta_sv": "Change in stroke volume",
        "hr": "Heart rate",
        "delta_rv_cpo": "Change in RV cardiac power output",
        "delta_hr": "Change in heart rate",
        "pre_pcwp": "Baseline PCWP",
        "pre_map": "Baseline MAP",
        "post_tdco": "Post-implant TDCO",
        "pre_egfr": "Baseline eGFR",
    }

    features = []
    for _, row in df.head(30).iterrows():
        name = str(row[first_col]).strip() if pd.notna(row[first_col]) else ""
        consensus_val = float(row[consensus_col]) if consensus_col and pd.notna(row.get(consensus_col)) else 0
        features.append({
            "name": name,
            "description": feature_descriptions.get(name, ""),
            "consensus": round(consensus_val, 4),
        })

    return {"features": features}


# ---------------------------------------------------------------------------
# Chart Generation (for HTML report)
# ---------------------------------------------------------------------------

def chart_to_base64(fig):
    """Convert matplotlib figure to base64 PNG."""
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    img = base64.b64encode(buf.read()).decode("utf-8")
    plt.close(fig)
    return img


def plot_survival_chart(survival_data: dict) -> str:
    """Bar chart of mortality by indication."""
    indications = survival_data.get("byIndication", [])
    if not indications:
        return ""

    labels = [d["indication"][:15] for d in indications]
    survived = [d["survived"] for d in indications]
    expired = [d["expired"] for d in indications]

    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(labels))
    w = 0.35
    bars1 = ax.bar(x - w / 2, survived, w, label="Survived", color="#34d399", alpha=0.8)
    bars2 = ax.bar(x + w / 2, expired, w, label="Expired", color="#f87171", alpha=0.8)

    for bar in bars1:
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                str(int(bar.get_height())), ha="center", va="bottom", fontsize=9, color="#34d399")
    for bar in bars2:
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                str(int(bar.get_height())), ha="center", va="bottom", fontsize=9, color="#f87171")

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    ax.set_ylabel("Patient Count")
    ax.set_title("Survival by Indication", fontsize=14, fontweight="bold")
    ax.legend()
    ax.set_facecolor("#f8f9fa")
    fig.tight_layout()
    return chart_to_base64(fig)


def plot_hemodynamic_chart(hemo_data: dict) -> str:
    """Bar chart of delta means with p-value annotations."""
    tests = hemo_data.get("statisticalTests", [])
    if not tests:
        return ""

    metrics = [t["metric"] for t in tests]
    surv_means = [t["survivedMean"] for t in tests]
    exp_means = [t["expiredMean"] for t in tests]
    pvals = [t["pValue"] for t in tests]

    fig, ax = plt.subplots(figsize=(12, 5))
    x = np.arange(len(metrics))
    w = 0.35
    ax.bar(x - w / 2, surv_means, w, label="Survived", color="#34d399", alpha=0.8)
    ax.bar(x + w / 2, exp_means, w, label="Expired", color="#f87171", alpha=0.8)

    for i, (m, p) in enumerate(zip(metrics, pvals)):
        sig = " *" if p < 0.05 else (" †" if p < 0.1 else "")
        ax.text(i, max(surv_means[i], exp_means[i]) + 0.02, f"p={p:.3f}{sig}",
                ha="center", va="bottom", fontsize=8, fontstyle="italic")

    ax.set_xticks(x)
    ax.set_xticklabels(metrics, rotation=20, ha="right")
    ax.set_ylabel("Mean Delta (Post − Pre)")
    ax.set_title("Hemodynamic Changes: Survived vs Expired", fontsize=14, fontweight="bold")
    ax.axhline(y=0, color="gray", linestyle="-", linewidth=0.5)
    ax.legend()
    ax.set_facecolor("#f8f9fa")
    fig.tight_layout()
    return chart_to_base64(fig)


def plot_responder_radar(responder_data: dict) -> str:
    """Radar-like bar chart comparing responder vs non-responder profiles."""
    profiles = responder_data.get("profileVariables", [])
    if not profiles:
        return ""

    # Take top 6 most significant variables
    profiles = sorted(profiles, key=lambda x: x["pValue"])[:8]
    metrics = [p["metric"][:12] for p in profiles]
    r_means = [p["responderMean"] for p in profiles]
    nr_means = [p["nonResponderMean"] for p in profiles]

    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(metrics))
    w = 0.35
    ax.bar(x - w / 2, r_means, w, label="Responders", color="#60a5fa", alpha=0.8)
    ax.bar(x + w / 2, nr_means, w, label="Non-Responders", color="#f87171", alpha=0.8)

    for i, p in enumerate(profiles):
        sig = " *" if p["significant"] else ""
        ax.text(i, max(r_means[i], nr_means[i]) + 0.02, f"p={p['pValue']:.3f}{sig}",
                ha="center", va="bottom", fontsize=7, fontstyle="italic", rotation=45)

    ax.set_xticks(x)
    ax.set_xticklabels(metrics, rotation=25, ha="right")
    ax.set_ylabel("Mean Value")
    ax.set_title("Responder vs Non-Responder: Baseline Characteristics", fontsize=14, fontweight="bold")
    ax.legend()
    ax.set_facecolor("#f8f9fa")
    fig.tight_layout()
    return chart_to_base64(fig)


def plot_pv_chart(pv_data: dict) -> str:
    """PV Loop metrics comparison."""
    tests = pv_data.get("statisticalTests", [])
    if not tests:
        return ""

    metrics = [t["metric"] for t in tests]
    surv_means = [t["survivedMean"] for t in tests]
    exp_means = [t["expiredMean"] for t in tests]

    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(metrics))
    w = 0.35
    ax.bar(x - w / 2, surv_means, w, label="Survived", color="#34d399", alpha=0.8)
    ax.bar(x + w / 2, exp_means, w, label="Expired", color="#f87171", alpha=0.8)

    for i, t in enumerate(tests):
        sig = " *" if t["significant"] else ""
        ax.text(i, max(surv_means[i], exp_means[i]) + 0.02, f"p={t['pValue']:.3f}{sig}",
                ha="center", va="bottom", fontsize=8, fontstyle="italic")

    ax.set_xticks(x)
    ax.set_xticklabels(metrics, rotation=20, ha="right")
    ax.set_ylabel("Value")
    ax.set_title("Ventricular Mechanics: Survived vs Expired", fontsize=14, fontweight="bold")
    ax.legend()
    ax.set_facecolor("#f8f9fa")
    fig.tight_layout()
    return chart_to_base64(fig)


def plot_lab_chart(lab_data: dict) -> str:
    """Lab recovery comparison chart."""
    tests = lab_data.get("statisticalTests", [])
    if not tests:
        return ""

    metrics = [t["metric"] for t in tests]
    surv_means = [t["survivedMean"] for t in tests]
    exp_means = [t["expiredMean"] for t in tests]

    fig, ax = plt.subplots(figsize=(12, 5))
    x = np.arange(len(metrics))
    w = 0.35
    ax.bar(x - w / 2, surv_means, w, label="Survived", color="#34d399", alpha=0.8)
    ax.bar(x + w / 2, exp_means, w, label="Expired", color="#f87171", alpha=0.8)

    for i, t in enumerate(tests):
        sig = " *" if t["significant"] else ""
        ax.text(i, max(surv_means[i], exp_means[i]) + 0.02, f"p={t['pValue']:.3f}{sig}",
                ha="center", va="bottom", fontsize=8, fontstyle="italic", rotation=45)

    ax.set_xticks(x)
    ax.set_xticklabels(metrics, rotation=25, ha="right")
    ax.set_ylabel("Mean Delta (Post − Pre)")
    ax.set_title("Lab Recovery: Survived vs Expired", fontsize=14, fontweight="bold")
    ax.axhline(y=0, color="gray", linestyle="-", linewidth=0.5)
    ax.legend()
    ax.set_facecolor("#f8f9fa")
    fig.tight_layout()
    return chart_to_base64(fig)


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------

def generate_html_report(survival: dict, hemo: dict, labs: dict, responders: dict,
                          pv: dict, ml_xref: dict) -> str:
    """Generate self-contained HTML report with inline charts."""
    survival_img = plot_survival_chart(survival)
    hemo_img = plot_hemodynamic_chart(hemo)
    lab_img = plot_lab_chart(labs)
    responder_img = plot_responder_radar(responders)
    pv_img = plot_pv_chart(pv)

    # Build statistical tests table
    all_tests = hemo.get("statisticalTests", []) + labs.get("statisticalTests", [])
    all_tests = sorted(all_tests, key=lambda x: x["pValue"])
    tests_rows = ""
    for t in all_tests[:15]:
        sig = "✓" if t["significant"] else ""
        tests_rows += f"""
        <tr>
            <td>{t['metric']}</td>
            <td>{t['survivedMean']}</td>
            <td>{t['expiredMean']}</td>
            <td>{t['pValue']}</td>
            <td>{sig}</td>
        </tr>"""

    # Build indication table
    ind_rows = ""
    for d in survival.get("byIndication", []):
        ind_rows += f"""
        <tr>
            <td>{d['indication']}</td>
            <td>{d['total']}</td>
            <td>{d['survived']}</td>
            <td>{d['expired']}</td>
            <td>{d['mortalityRate']}%</td>
        </tr>"""

    # Build responder profile table
    resp_rows = ""
    for p in responders.get("profileVariables", [])[:8]:
        sig = "✓" if p["significant"] else ""
        resp_rows += f"""
        <tr>
            <td>{p['metric']}</td>
            <td>{p['responderMean']}</td>
            <td>{p['nonResponderMean']}</td>
            <td>{p['pValue']}</td>
            <td>{sig}</td>
        </tr>"""

    # Build ML features table
    ml_rows = ""
    for f in ml_xref.get("features", [])[:15]:
        ml_rows += f"""
        <tr>
            <td>{f['name']}</td>
            <td>{f.get('description', '')}</td>
            <td>{f.get('consensus', '')}</td>
        </tr>"""

    overall = survival.get("overall", {})
    responder_summary = responders.get("summary", {})

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Impella Effectiveness Analysis Report</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0f172a; color: #e2e8f0; line-height: 1.6; }}
    .container {{ max-width: 1000px; margin: 0 auto; padding: 40px 20px; }}
    h1 {{ font-size: 2rem; font-weight: 300; margin-bottom: 8px; color: #c084fc; }}
    h2 {{ font-size: 1.4rem; font-weight: 600; margin: 40px 0 16px; color: #e2e8f0; border-bottom: 1px solid #334155; padding-bottom: 8px; }}
    h3 {{ font-size: 1.1rem; font-weight: 500; margin: 24px 0 12px; color: #94a3b8; }}
    .subtitle {{ color: #64748b; font-size: 0.9rem; margin-bottom: 24px; }}
    .summary-cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin: 24px 0; }}
    .card {{ background: #1e293b; border: 1px solid #334155; border-radius: 12px; padding: 20px; }}
    .card .label {{ font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.1em; color: #64748b; }}
    .card .value {{ font-size: 1.8rem; font-weight: 300; margin: 4px 0; }}
    .card .note {{ font-size: 0.8rem; color: #64748b; }}
    .chart {{ margin: 24px 0; background: #1e293b; border-radius: 12px; padding: 20px; border: 1px solid #334155; }}
    .chart img {{ width: 100%; height: auto; border-radius: 8px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 0.85rem; }}
    th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid #1e293b; }}
    th {{ background: #1e293b; color: #94a3b8; font-weight: 600; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em; }}
    tr:hover {{ background: #1e293b; }}
    .sig {{ color: #34d399; }}
    .p-footer {{ margin-top: 60px; padding-top: 20px; border-top: 1px solid #334155; font-size: 0.8rem; color: #64748b; text-align: center; }}
    .section {{ margin: 32px 0; }}
    @media print {{ body {{ background: white; color: #1e293b; }} .card {{ border-color: #ddd; }} .chart {{ border-color: #ddd; }} th {{ background: #f1f5f9; }} }}
</style>
</head>
<body>
<div class="container">

<h1>Impella Effectiveness Analysis</h1>
<p class="subtitle">Comprehensive multi-dimensional analysis of Impella hemodynamic support outcomes</p>

<div class="summary-cards">
    <div class="card">
        <div class="label">Total Patients</div>
        <div class="value">{overall.get('total', 0)}</div>
        <div class="note">In analysis cohort</div>
    </div>
    <div class="card">
        <div class="label">Mortality Rate</div>
        <div class="value" style="color: #f87171">{overall.get('mortalityRate', 0)}%</div>
        <div class="note">Overall cohort</div>
    </div>
    <div class="card">
        <div class="label">Responder Rate</div>
        <div class="value" style="color: #60a5fa">{responder_summary.get('responderRate', 0)}%</div>
        <div class="note">CPO↑ + Lactate↓ + Survived</div>
    </div>
    <div class="card">
        <div class="label">Survived</div>
        <div class="value" style="color: #34d399">{overall.get('survived', 0)}</div>
        <div class="note">vs {overall.get('expired', 0)} expired</div>
    </div>
</div>

<div class="section">
<h2>1. Survival & Outcomes</h2>
<h3>Mortality by Indication</h3>
<div class="chart"><img src="data:image/png;base64,{survival_img}" alt="Survival by Indication"></div>
<table>
    <tr><th>Indication</th><th>Total</th><th>Survived</th><th>Expired</th><th>Mortality</th></tr>
    {ind_rows}
</table>
</div>

<div class="section">
<h2>2. Hemodynamic Response</h2>
<h3>Delta Analysis: Survived vs Expired</h3>
<p class="subtitle">Mean change (48h post − pre-implant). Asterisk (*) = p &lt; 0.05, Dagger (†) = p &lt; 0.10.</p>
<div class="chart"><img src="data:image/png;base64,{hemo_img}" alt="Hemodynamic Changes"></div>
</div>

<div class="section">
<h2>3. Organ Function & Labs</h2>
<div class="chart"><img src="data:image/png;base64,{lab_img}" alt="Lab Recovery"></div>
</div>

<div class="section">
<h2>4. Responder Profiling</h2>
<p class="subtitle">Responder defined as: CPO ↑ AND Lactate ↓ AND Survived. Chart compares baseline characteristics.</p>
<div class="chart"><img src="data:image/png;base64,{responder_img}" alt="Responder Profiles"></div>
<table>
    <tr><th>Metric</th><th>Responder Mean</th><th>Non-Responder Mean</th><th>p-Value</th><th>Sig</th></tr>
    {resp_rows}
</table>
</div>

<div class="section">
<h2>5. Ventricular Mechanics (PV Loop)</h2>
<div class="chart"><img src="data:image/png;base64,{pv_img}" alt="Ventricular Mechanics"></div>
</div>

<div class="section">
<h2>6. ML Model Cross-Reference</h2>
<p class="subtitle">Top consensus features from mortality prediction models, mapped to clinical mechanisms.</p>
<table>
    <tr><th>Feature</th><th>Description</th><th>Consensus Score</th></tr>
    {ml_rows}
</table>
</div>

<div class="section">
<h2>7. Statistical Test Summary</h2>
<p class="subtitle">Mann-Whitney U tests comparing delta values between survived and expired groups (top 15 by significance).</p>
<table>
    <tr><th>Metric</th><th>Survived Mean</th><th>Expired Mean</th><th>p-Value</th><th>Sig</th></tr>
    {tests_rows}
</table>
</div>

<div class="section">
<h2>8. Limitations</h2>
<ul style="color: #94a3b8; margin-left: 20px;">
    <li>Sample size: 67 patients with detailed hemodynamics, 112 in cohort — limited generalizability</li>
    <li>Missing data: ~30-50% for echo and PV loop variables</li>
    <li>Single-center retrospective analysis</li>
    <li>No adjustment for confounders (illness severity, comorbidities)</li>
    <li>Responder definition (CPO↑ + lactate↓ + survived) is one of many possible definitions</li>
    <li>AMI/CGS subgroup (N=10) may drive some statistical findings</li>
</ul>
</div>

<div class="p-footer">
    Generated by Impella Analytics &mdash; {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}
</div>

</div>
</body>
</html>"""
    return html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def clean_nan(obj):
    """Recursively replace NaN/Inf with None and convert numpy types."""
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return None if np.isnan(obj) or np.isinf(obj) else float(obj)
    if isinstance(obj, float) and (np.isnan(obj) or np.isinf(obj)):
        return None
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [clean_nan(v) for v in obj]
    return obj


def main():
    print("=" * 60)
    print("Impella Effectiveness Analysis")
    print("=" * 60)

    print("\n[1/4] Loading data...")
    df = load_patient_data(DATA_PATH)
    cohort = load_cohort(DATA_PATH)
    print(f"  Patient Data: {len(df)} patients")
    print(f"  Cohort: {len(cohort)} patients")

    print("\n[2/4] Computing analyses...")
    survival = compute_survival_analysis(df, cohort)
    print(f"  Survival: {survival['overall']['survived']} survived / {survival['overall']['expired']} expired")

    hemo = compute_hemodynamic_analysis(df, cohort)
    sig_hemo = [t for t in hemo["statisticalTests"] if t["significant"]]
    print(f"  Hemodynamics: {len(hemo['deltas'])} data points, {len(sig_hemo)} significant tests")

    labs = compute_lab_analysis(df, cohort)
    sig_labs = [t for t in labs["statisticalTests"] if t["significant"]]
    print(f"  Labs: {len(sig_labs)} significant tests")

    responders = compute_responder_profiles(df, cohort)
    print(f"  Responders: {responders['summary']['totalResponders']} / {responders['summary']['totalResponders'] + responders['summary']['totalNonResponders']} ({responders['summary']['responderRate']}%)")

    pv = compute_ventricular_mechanics(df, cohort)
    print(f"  PV Loop: {len(pv['values'])} data points")

    ml_xref = compute_ml_cross_reference()
    print(f"  ML Features: {len(ml_xref['features'])} consensus features")

    print("\n[3/4] Exporting JSON...")
    output = {
        "survival": survival,
        "hemodynamics": hemo,
        "labs": labs,
        "responders": responders,
        "ventricularMechanics": pv,
        "mlCrossReference": ml_xref,
        "generatedAt": pd.Timestamp.now().isoformat(),
    }
    output = clean_nan(output)
    with open(OUTPUT_JSON, "w") as f:
        json.dump(output, f, indent=2)
    print(f"  Written: {OUTPUT_JSON}")

    print("\n[4/4] Generating HTML report...")
    html = generate_html_report(survival, hemo, labs, responders, pv, ml_xref)
    with open(OUTPUT_REPORT, "w") as f:
        f.write(html)
    print(f"  Written: {OUTPUT_REPORT}")

    print("\nAnalysis complete.")
    print(f"  Dashboard data: {OUTPUT_JSON}")
    print(f"  Report:         {OUTPUT_REPORT}")


if __name__ == "__main__":
    main()
